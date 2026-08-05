"""Ingest ADI parametric-search exports (.xlsx) into partdb.

Analog Devices refuses scripted access to analog.com (every datasheet URL returns
HTTP 403, and the product-category pages 403 as well), so there is no polite way
to crawl them. There is no need to: their parametric search exports exactly the
table we want, and one export per product family covers the whole RF catalogue.

    Downloads/ADIParametrics/
        ADIParametricSearchAmplifiers.xlsx
        ADIParametricSearchLNA+PA.xlsx
        ADIParametricSearchMixers.xlsx
        ... etc

WORKBOOK SHAPE (verified against the real exports)
    sheet 'Cover'            row 1  Product Type -> the category
    sheet 'Web Display'      row 1  column headers
                             row 2  units row  (Hz, dB, dBm, dBc/Hz ...)
                             row 3+ one part per row
    Part Number is an Excel formula, NOT a plain value:
        =HYPERLINK("https://www.analog.com/en/ADL6346B#details", "ADL6346B")
    so the workbook must be read with data_only=False and the formula parsed --
    with data_only=True that column comes back empty and every row looks blank.
    The formula conveniently carries the product URL as well as the part number.

Frequencies are in Hz in these exports (1500000000), so they are scaled to GHz.

Run standalone:
    python -m rfparts.adi_parametric_ingest "C:\\path\\ADIParametrics" --dry-run
"""
from __future__ import annotations

import argparse
import re
from collections import Counter
from pathlib import Path

try:
    from .partdb import SpecRow, upsert_part, put_specs, put_evidence
except ImportError:                                     # loose-script fallback
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from rfparts.partdb import (                          # type: ignore
        SpecRow, upsert_part, put_specs, put_evidence)

VENDOR = "Analog Devices"
DATASHEET_TPL = ("https://www.analog.com/media/en/technical-documentation/"
                 "data-sheets/{pn}.pdf")

# 'Product Type' from the Cover sheet -> canonical pipeline category.
_PRODUCT_TYPE_CAT = [
    ("low-noise amplifier", "amplifier"), ("power amplifier", "amplifier"),
    ("rf amplifier", "amplifier"), ("amplifier", "amplifier"),
    ("attenuator", "attenuator"),
    ("beamformer", "beamformer"), ("vector modulator", "beamformer"),
    ("mixer", "mixer"),
    ("modulator", "modulator"), ("demodulator", "modulator"),
    ("phase shifter", "phase_shifter"),
    ("switch", "switch"),
    ("tunable filter", "filter"), ("filter", "filter"),
    ("detector", "detector"),
    ("synthesizer", "synthesizer"), ("pll", "synthesizer"),
]

# Per-row 'type' column (e.g. 'RF Amp Type') that refines the category.
_ROW_TYPE_SUB = [
    ("low noise", ("amplifier", "lna")), ("lna", ("amplifier", "lna")),
    ("power amplifier", ("amplifier", "pa")),
    ("gain block", ("amplifier", "buffer")),
    ("driver", ("amplifier", "driver")),
    ("variable gain", ("amplifier", "vga")),
    ("distributed", ("amplifier", "")),
]

# Column header (lowercased, fuzzy) -> (partdb spec key, unit, how to read).
# 'range' means the header names a min/max pair handled separately.
# Bump when the PARSING changes so a rebuild re-reads an unchanged workbook once.
# Without this the frequency-unit fix below could not reach parts already in the
# database: the file has not changed, so the source checkpoint skipped it and the
# wrong values simply stayed.
PARSER_VERSION = 3

_COL_SPECS = [
    # ---- frequency. The IF/LO variants MUST precede the generic RF ones: a
    # mixer sheet carries "Frequency Response IF min" alongside "Frequency
    # Response min", and dropping the qualified ones lost the IF and LO ranges
    # entirely.
    (r"frequency response if min", "if_freq_min", "GHz"),
    (r"frequency response if max", "if_freq_max", "GHz"),
    (r"frequency response lo min", "lo_freq_min", "GHz"),
    (r"frequency response lo max", "lo_freq_max", "GHz"),
    (r"frequency response rf min", "freq_min", "GHz"),
    (r"frequency response rf max", "freq_max", "GHz"),
    (r"frequency response min|freq(uency)? min|min frequency", "freq_min", "GHz"),
    (r"frequency response max|freq(uency)? max|max frequency", "freq_max", "GHz"),
    (r"specified at frequency", "specified_at_ghz", "GHz"),
    (r"bandwidth\s*-?\s*3\s*db|^bandwidth", "bandwidth_ghz", "GHz"),
    (r"frequency cut+off range|cutoff range", "cutoff_range_ghz", "GHz"),
    # ---- gain / noise / linearity
    (r"noise figure", "nf_db", "dB"),
    (r"\biip3\b|input ip3", "iip3_dbm", "dBm"),
    (r"\boip3\b|output ip3", "oip3_dbm", "dBm"),
    (r"ip-?0\.1\s*db|ip0\.1db|p0\.1db", "p01db_dbm", "dBm"),
    (r"op1db|ip1db|p1db|output p1db", "p1db_dbm", "dBm"),
    (r"^gain( typ)?$|gain typ|small signal gain", "gain_db", "dB"),
    (r"isolation", "isolation_db", "dB"),
    (r"insertion loss", "insertion_loss_db", "dB"),
    (r"return loss", "return_loss_db", "dB"),
    # ADI's column is "Conversion Gain" in dB with a MEANINGFUL sign: +10 for an
    # active mixer, -10 for a passive one (i.e. 10 dB of loss). Filing that
    # straight into conversion_loss_db made a passive mixer look like -10 dB of
    # loss, which passes any "loss below N dB" filter. Stored as gain, with the
    # loss derived below.
    (r"conversion gain", "conversion_gain_db", "dB"),
    (r"conversion loss", "conversion_loss_db", "dB"),
    (r"attenuation range|attenuation", "attenuation_db", "dB"),
    (r"residual phase noise", "phase_noise_dbc", "dBc/Hz"),
    # ---- power / speed
    (r"max average rf power|max rf power", "max_rf_power_dbm", "dBm"),
    (r"output power|psat|pout", "psat_dbm", "dBm"),
    (r"switching speed", "switching_speed_ns", "ns"),
    (r"phase adjust resolution|phase shifter resolution", "phase_resolution",
     "deg"),
    # ---- packaging and commercial. 'Package' was unmapped in eight of the nine
    # workbooks, so package data never reached the database at all.
    (r"^package$|package type", "package", ""),
    (r"package area", "package_area_mm2", "mm2"),
    (r"package height", "package_height_mm", "mm"),
    (r"^availability", "availability", ""),
    (r"1ku list price|price", "price_usd", "USD"),
]

# Frequency-like keys are stored in Hz in these exports and must be scaled.
_FREQ_KEYS = {"freq_min", "freq_max", "if_freq_min", "if_freq_max",
              "lo_freq_min", "lo_freq_max", "specified_at_ghz",
              "bandwidth_ghz", "cutoff_range_ghz"}
# Text columns: keep the string rather than pulling the first number out of it.
_TEXT_KEYS = {"package", "availability"}
_HYPERLINK = re.compile(r'HYPERLINK\(\s*"([^"]+)"\s*,\s*"([^"]+)"\s*\)', re.I)
_TEMP = re.compile(r"(-?\d+)\s*(?:to|\.\.|–|-)\s*\+?(-?\d+)\s*°?\s*C", re.I)


# Roughly 15% of rows in these exports have the literal string 'None' in every
# parametric column while the DESCRIPTION still states the range --
# "1GHz to 22GHz, 15W, GaN Power Amplifier". Reading it back out recovers RF
# specs for parts that would otherwise land in the table with nothing at all.
_DESC_RANGE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(GHz|MHz|kHz)?\s*(?:to|-|–|through)\s*"
    r"(\d+(?:\.\d+)?)\s*(GHz|MHz|kHz)", re.I)
_DESC_SINGLE = re.compile(r"\b(?:up to|dc\s*(?:to|-)\s*)(\d+(?:\.\d+)?)\s*"
                          r"(GHz|MHz)", re.I)
_DESC_POWER_W = re.compile(r"(\d+(?:\.\d+)?)\s*W\b", re.I)
# "45 dB Linear-in-dB Variable Gain Amplifier" -- the number precedes the word by
# several tokens, so requiring "dB gain" adjacent missed it. Only 1 of 234 space
# portfolio parts states a gain at all, but it is free to take.
_DESC_GAIN = re.compile(r"(\d+(?:\.\d+)?)\s*dB\b(?=[^.]{0,40}?\bgain\b)", re.I)
_DESC_GAIN_RANGE = re.compile(r"variable\s+gain|linear-in-db|gain\s+control", re.I)
_DESC_NF = re.compile(r"(\d+(?:\.\d+)?)\s*dB\s*(?:noise figure|nf)", re.I)
_SCALE = {"ghz": 1.0, "mhz": 1e-3, "khz": 1e-6}


def specs_from_description(desc):
    """Best-effort RF specs from an ADI description string."""
    out = {}
    if not desc:
        return out
    m = _DESC_RANGE.search(desc)
    if m:
        lo, lo_u, hi, hi_u = m.group(1), m.group(2), m.group(3), m.group(4)
        unit_hi = (hi_u or "GHz").lower()
        unit_lo = (lo_u or hi_u or "GHz").lower()
        a = float(lo) * _SCALE.get(unit_lo, 1.0)
        b = float(hi) * _SCALE.get(unit_hi, 1.0)
        out["freq_min"] = (min(a, b), "GHz")
        out["freq_max"] = (max(a, b), "GHz")
    else:
        m = _DESC_SINGLE.search(desc)
        if m:
            b = float(m.group(1)) * _SCALE.get(m.group(2).lower(), 1.0)
            out["freq_min"] = (0.0, "GHz")
            out["freq_max"] = (b, "GHz")
    m_dbm = re.search(r"(\d+(?:\.\d+)?)\s*dBm", desc, re.I)
    if m_dbm:
        out["psat_dbm"] = (float(m_dbm.group(1)), "dBm")
    m = None if m_dbm else _DESC_POWER_W.search(desc)
    if m:
        watts = float(m.group(1))
        if watts > 0:
            import math
            out["psat_dbm"] = (round(10 * math.log10(watts * 1000.0), 1), "dBm")
    for rx, key, unit in ((_DESC_GAIN, "gain_db", "dB"),
                          (_DESC_NF, "nf_db", "dB")):
        m = rx.search(desc)
        if m:
            # On a VGA the figure is a gain CONTROL RANGE, not a gain. Recording it
            # as gain_db would flatter the part in any "gain above N" filter.
            if key == "gain_db" and _DESC_GAIN_RANGE.search(desc):
                out["gain_range_db"] = (float(m.group(1)), unit)
            else:
                out[key] = (float(m.group(1)), unit)
    return out


def _num(v):
    if v is None:
        return None
    if isinstance(v, str) and v.strip().lower() in ("none", "n/a", "na", "-", ""):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    m = re.search(r"-?\d+(?:\.\d+)?", str(v).replace(",", ""))
    return float(m.group(0)) if m else None


def _category_for(product_type):
    low = (product_type or "").lower()
    for needle, cat in _PRODUCT_TYPE_CAT:
        if needle in low:
            return cat
    return "ic"


def _refine(category, row_type):
    low = (row_type or "").lower()
    for needle, (cat, sub) in _ROW_TYPE_SUB:
        if needle in low:
            return cat, sub
    return category, ""


_HEADER_FREQ_UNIT = re.compile(r"\(\s*(g|m|k)?hz\s*\)", re.I)
_FREQ_UNIT_SCALE = {"g": 1.0, "m": 1e-3, "k": 1e-6, "": 1e-9}   # -> GHz


def freq_unit_from_header(header):
    """GHz-scale factor stated in a column header, or None if it says nothing.

    ADI's exports label the unit in the header -- "Frequency Response RF Min
    (MHz)", "(Hz)", "(GHz)" -- and that label was being thrown away in favour of
    guessing from the magnitude. See _freq_to_ghz for why the guess cannot work.
    """
    m = _HEADER_FREQ_UNIT.search(str(header or ""))
    if not m:
        return None
    return _FREQ_UNIT_SCALE.get((m.group(1) or "").lower())


# The cell itself often carries the unit -- ADI's switch export writes
# "100 MHz", "6 GHz", "4 GHz" as TEXT, with no unit in the column header at all.
# That is the most authoritative signal available and it was being ignored: _num()
# pulled out the bare 100, the header offered nothing, and the magnitude fallback
# read it as 100 GHz. HMC8038 ended up with a 6-100 GHz band when its own
# description says 0.1 GHz to 6.0 GHz.
_CELL_FREQ_UNIT = re.compile(
    r"(-?\d[\d,]*(?:\.\d+)?)\s*(thz|ghz|mhz|khz|hz)\b", re.I)
_CELL_UNIT_SCALE = {"thz": 1e3, "ghz": 1.0, "mhz": 1e-3, "khz": 1e-6, "hz": 1e-9}


def freq_from_cell(value):
    """(number, GHz-scale) when the CELL states its own unit, else None.

    Checked before the column header and before any magnitude guess, because a
    unit written next to the number is the vendor telling you outright."""
    if value is None:
        return None
    m = _CELL_FREQ_UNIT.search(str(value))
    if not m:
        return None
    try:
        n = float(m.group(1).replace(",", ""))
    except ValueError:
        return None
    return n, _CELL_UNIT_SCALE[m.group(2).lower()]


def _freq_to_ghz(v, scale=None, plausible=(1e-6, 300.0)):
    """A frequency column value in GHz.

    The unit comes from the column HEADER when it states one. It has to: the old
    code inferred the unit from the size of the number --

        n > 1e6  -> Hz ;  n > 1e3 -> MHz ;  otherwise assume GHz

    -- and that is not decidable. 2700 MHz landed above the 1e3 threshold and was
    correctly scaled to 2.7 GHz, while 450 MHz fell through every branch and was
    stored as 450 GHz, in the same part. 1000 MHz failed too, the threshold being
    exclusive. Anything in 1..1000 is genuinely ambiguous from magnitude alone.

    With no stated unit the magnitude fallback is kept -- it is the only signal
    left -- but the result is checked against a plausible RF range and dropped
    rather than stored if it lands outside, because a wrong frequency silently
    poisons every band filter the part appears in.
    """
    # 1. the cell's own unit wins over everything else
    cell = freq_from_cell(v)
    if cell is not None:
        n, cscale = cell
        g = n * cscale
        return g if g >= 0 else None
    n = _num(v)
    if n is None:
        return None
    if scale is not None:
        # The header stated the unit: honour it, including values that look odd.
        # Second-guessing the file is how the original bug happened.
        g = n * scale
        # 0 is a legitimate lower edge: DC-coupled parts really do start at DC
        # (HMC8038 is DC - 6 GHz). Rejecting it threw away the min of every such
        # part and left a one-sided band.
        return g if g >= 0 else None
    # No stated unit, so the magnitude is the only signal. Calibrated against what
    # an RF part can actually be rather than round numbers: nothing in these
    # exports operates at 450 GHz, so a bare 450 must be MHz. The GHz reading is
    # only kept where it is physically credible.
    lo, hi = plausible
    if n > 1e6:
        g = n / 1e9                 # certainly Hz
    elif n >= 1e3:
        g = n / 1e3                 # certainly MHz
    elif n > hi:
        g = n / 1e3                 # too big to be GHz -> MHz
    else:
        g = n                       # credible as GHz
    if g < 0 or g > hi:
        return None
    if g > 0 and g < lo:
        return None
    return g                        # 0 == DC, kept


def _hz_to_ghz(v):                  # kept for callers outside this module
    return _freq_to_ghz(v)


def parse_workbook(path, verbose=False):
    """Yield one dict per part row."""
    import openpyxl
    path = Path(path)
    # data_only=False is essential: the part number is a HYPERLINK formula.
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)

    product_type = ""
    if "Cover" in wb.sheetnames:
        for row in wb["Cover"].iter_rows(max_row=12, values_only=True):
            cells = [str(c) for c in row if c is not None]
            if len(cells) >= 2 and "product type" in cells[0].lower():
                product_type = cells[1]
                break
    sheet = ("Web Display" if "Web Display" in wb.sheetnames
             else wb.sheetnames[-1])
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return
    headers = [str(h or "").strip() for h in rows[0]]
    # row index 1 is usually the units row; detect it rather than assume
    body_from = 1
    if len(rows) > 1:
        second = [str(c or "") for c in rows[1]]
        if sum(1 for c in second if c and len(c) <= 12) >= 3 and \
                not any(_HYPERLINK.search(c) for c in second):
            body_from = 2
    category = _category_for(product_type)
    if verbose:
        print(f"      product type '{product_type}' -> category '{category}'; "
              f"{len(headers)} column(s), data starts at row {body_from + 1}")

    # map columns
    col_map, type_col = {}, None
    for i, h in enumerate(headers):
        hl = h.lower()
        if "type" in hl and i > 0 and not col_map.get(i):
            type_col = type_col if type_col is not None else i
        for pat, key, unit in _COL_SPECS:
            if re.search(pat, hl):
                # Remember the scale the HEADER states, so frequency columns are
                # converted from what the file says rather than from a guess.
                col_map[i] = (key, unit, freq_unit_from_header(h))
                break
        if "temperature" in hl:
            col_map[i] = ("temp_range", "C", None)
        elif "lifecycle" in hl:
            col_map[i] = ("lifecycle", "", None)
        elif hl.startswith("description"):
            col_map[i] = ("description", "", None)

    for r in rows[body_from:]:
        if not r:
            continue
        cell0 = str(r[0]) if r[0] is not None else ""
        m = _HYPERLINK.search(cell0)
        if m:
            product_url, pn = m.group(1), m.group(2).strip()
        else:
            pn, product_url = cell0.strip(), ""
        if not pn or len(pn) < 3:
            continue
        rec = {"mpn": pn, "product_url": product_url, "category": category,
               "subcategory": "", "description": "", "specs": {},
               "source_family": product_type}
        row_type = (str(r[type_col]) if type_col is not None
                    and type_col < len(r) and r[type_col] is not None else "")
        rec["category"], rec["subcategory"] = _refine(category, row_type)
        if row_type and row_type.strip().lower() not in ("none", "nan"):
            # 'RF Amp Type', 'Attenuator Type', 'Configuration', ... -- the
            # vendor's own classification, worth keeping verbatim alongside our
            # normalised subcategory.
            rec["specs"]["device_type"] = (row_type.strip()[:60], "")
        for i, (key, unit, fscale) in col_map.items():
            if i >= len(r) or r[i] is None:
                continue
            val = r[i]
            if key == "description":
                rec["description"] = str(val)[:200]
            elif key == "temp_range":
                tm = _TEMP.search(str(val))
                if tm:
                    rec["specs"]["temp_min_c"] = (float(tm.group(1)), "C")
                    rec["specs"]["temp_max_c"] = (float(tm.group(2)), "C")
            elif key == "lifecycle":
                rec["specs"]["lifecycle"] = (str(val)[:60], "")
            elif key in _FREQ_KEYS:
                g = _freq_to_ghz(val, fscale)
                if g is not None:
                    rec["specs"][key] = (g, "GHz")
            elif key in _TEXT_KEYS:
                txt = str(val).strip()
                if txt and txt.lower() not in ("nan", "none", "-"):
                    rec["specs"][key] = (txt[:60], "")
            else:
                n = _num(val)
                if n is not None:
                    rec["specs"][key] = (n, unit)
                    if key == "conversion_gain_db" and n < 0:
                        # negative gain is loss; publish both so either filter
                        # behaves sensibly
                        rec["specs"]["conversion_loss_db"] = (abs(n), "dB")
        # fill gaps from the description, never overwriting a real column value
        if rec.get("description"):
            for key, val in specs_from_description(rec["description"]).items():
                if key in rec["specs"]:
                    continue          # the sheet already supplied a real value
                rec["specs"][key] = val
                rec.setdefault("spec_source", {})[key] = "description"
        yield rec


# These exports are ADI's general RF catalogue, not their space portfolio. A part
# here may well also be listed as space qualified, but that fact comes from
# adi_space_ingest (or everythingRF), never from this file. Before classify_file
# distinguished them, every .xlsx was routed to the legacy adi adapter, which
# stamps space_variant on every row -- so the whole parametric catalogue was being
# reported as space qualified.
_NEVER_WRITE = {"space", "space_variant", "erf_grade"}


def _write(rec):
    pn = rec["mpn"]
    ds = DATASHEET_TPL.format(pn=pn.lower())
    pid = upsert_part(mpn=pn, vendor=VENDOR, category=rec["category"],
                      subcategory=rec.get("subcategory", ""),
                      product_url=rec.get("product_url") or ds,
                      description=rec.get("description", "")[:200])
    rows = []
    s = rec["specs"]
    lo, hi = s.get("freq_min"), s.get("freq_max")
    if lo or hi:
        a = (lo or hi)[0]
        b = (hi or lo)[0]
        rows.append(SpecRow(key="freq_ghz", value_min=min(a, b),
                            value_max=max(a, b), unit="GHz", method="catalog",
                            confidence=0.9, source_url=rec.get("product_url", ""),
                            snippet="ADI parametric export"))
    for key, (val, unit) in s.items():
        if key in ("freq_min", "freq_max"):
            continue
        if isinstance(val, str):
            rows.append(SpecRow(key=key, value_text=val, unit=unit,
                                method="catalog", confidence=0.9,
                                source_url=rec.get("product_url", "")))
        else:
            rows.append(SpecRow(key=key, value_typ=val, unit=unit,
                                method="catalog", confidence=0.9,
                                source_url=rec.get("product_url", "")))
    rows.append(SpecRow(key="datasheet_url", value_text=ds, method="catalog",
                        confidence=0.9, source_url=rec.get("product_url", "")))
    rows = [r for r in rows if r.key not in _NEVER_WRITE]
    put_specs(pid, rows)
    put_evidence(pid, [("adi-parametric-export", 3.0,
                        rec.get("product_url", ""),
                        f"ADI {rec.get('source_family', '')} parametric export")])
    return pid


def ingest(path, dry_run=False, verbose=False, progress=None, categories=None,
           part=None, cancel=None):
    """`path` may be one .xlsx or a folder of them."""
    def emit(m):
        (progress or print)(m)

    def stopped():
        try:
            if cancel is None:
                return False
            return bool(cancel.is_set() if hasattr(cancel, "is_set")
                        else cancel())
        except Exception:
            return False

    path = Path(path)
    files = ([path] if path.is_file()
             else sorted(p for p in path.glob("*.xlsx")
                         if not p.name.startswith("~$")))
    if not files:
        raise SystemExit(f"no .xlsx found at {path}")
    counts = Counter()
    seen = set()
    allowed = [c for c in (categories or []) if c]
    for f in files:
        if stopped():
            emit("    stop requested; leaving ADI ingest")
            break
        emit(f"    {f.name}")
        n = 0
        for rec in parse_workbook(f, verbose=verbose):
            if allowed and rec["category"] not in allowed:
                counts["filtered_out"] += 1
                continue
            key = rec["mpn"].upper()
            if key in seen:
                counts["duplicate_rows"] += 1
                continue
            seen.add(key)
            n += 1
            counts["parts"] += 1
            counts[f"cat:{rec['category']}"] += 1
            if verbose:
                sp = ", ".join(f"{k}={v[0]}" for k, v in
                               list(rec["specs"].items())[:4])
                emit(f"        {rec['mpn']:<18} {rec['category']:<12} {sp}")
            if part:
                # same dict shape the vendor walkers report, so the live table
                # does not need to know which source a row came from
                part({
                    "vendor": VENDOR, "mpn": rec["mpn"],
                    "category": rec["category"],
                    "subcategory": rec.get("subcategory", ""),
                    "specs": {k: (v[0] if isinstance(v, (tuple, list)) and v
                                  else v) for k, v in rec["specs"].items()},
                    "space": "", "url": rec.get("product_url", ""),
                    "source": f"ADI {rec.get('source_family', '')}",
                    "datasheet_url": DATASHEET_TPL.format(pn=rec["mpn"].lower()),
                })
            if not dry_run:
                _write(rec)
        emit(f"      -> {n} part(s)")
    return dict(counts)


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("path", help="an ADI parametric .xlsx, or a folder of them")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--verbose", action="store_true")
    args = ap.parse_args(argv)
    c = ingest(Path(args.path), args.dry_run, args.verbose)
    print(f"\n=== ADI parametric ingest {'(dry run) ' if args.dry_run else ''}===")
    print(f"  parts           {c.get('parts', 0)}")
    print(f"  duplicate rows  {c.get('duplicate_rows', 0)}")
    for k, v in sorted(c.items()):
        if k.startswith("cat:"):
            print(f"    {k[4:]:<14} {v}")


if __name__ == "__main__":
    main()
