#!/usr/bin/env python3
"""spacequal_catalog_debug — can we actually parse vendor catalogs and datasheets?

Nothing is committed anywhere. This answers three questions, per vendor, and
writes a report:

    1. CAN WE FETCH a catalog/listing page at all?
    2. CAN WE PARSE part numbers and datasheet URLs out of it?
    3. CAN WE FETCH AND READ one of those datasheet PDFs?

Only if all three pass for a vendor is it worth wiring into the downloader.

    python spacequal_catalog_debug.py                     # probe + parse + 2 PDFs
    python spacequal_catalog_debug.py --vendor qorvo      # one vendor
    python spacequal_catalog_debug.py --local saved_pages # parse pages YOU saved
    python spacequal_catalog_debug.py --no-pdf            # skip the PDF stage

RESEARCH BEHIND THE URL PATTERNS
--------------------------------
QORVO -- solved, and better than expected. The listing page
    https://www.qorvo.com/products/product-list?categoryID=<id>
and the category pages under /products/<family>/<subfamily> are SERVER-RENDERED
parametric tables, and each row carries BOTH links:
    <a href="/products/p/QPF4200">QPF4200</a> ... <a href="/products/d/da006470">
So the opaque datasheet id sits next to its part number and no per-product page
visit is needed. Category tree taken from Qorvo's own nav menu (below).

MARKI -- category pages are
    https://markimicrowave.com/products/<form>/<category>/
with form in {connectorized, surface-mount, bare-die, waveguide}. Product pages
are .../<category>/<pn>/ with a .../<pn>/datasheet/ child. Older PDFs also sit at
    https://markimicrowave.com/Assets/datasheets/<PN>.pdf
while newer ones are behind a UUID, which is why parsing the page matters.

MINI-CIRCUITS -- already solved by the local products JSON; included here as the
control case, since we know its PDF chain works end to end.

ADI / MACOM / SKYWORKS -- their DATASHEET url patterns are known and confirmed
(analog.com/media/.../data-sheets/<pn>.pdf, cdn.macom.com/datasheets/<PN>.pdf,
skyworksinc.com/-/media/.../<pn>-datasheet.pdf), but their LISTING pages are not
yet reverse-engineered. Candidate URLs are probed here so the report tells us
which are server-rendered and worth parsing, rather than me guessing.

Every request has a hard timeout and runs under a watchdog, so this cannot hang.
"""
from __future__ import annotations

import argparse
import io
import os
import re
import socket
import sys
import threading
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
from collections import OrderedDict
from pathlib import Path

socket.setdefaulttimeout(20)
UA = "rfparts-spacequal-debug/1.1 (catalog parser diagnostic)"
OUT = io.StringIO()


def say(line=""):
    print(line, flush=True)
    OUT.write(line + "\n")


def guarded(label, fn, budget):
    box = {}
    t0 = time.time()

    def target():
        try:
            box["v"] = fn()
            box["s"] = "ok"
        except Exception as e:
            box["v"] = f"{type(e).__name__}: {e}"
            box["s"] = "error"
    th = threading.Thread(target=target, daemon=True, name=label)
    th.start()
    th.join(budget)
    if th.is_alive():
        return "TIMEOUT", f"still blocked after {budget}s", time.time() - t0
    return box.get("s", "error"), box.get("v"), time.time() - t0


def http_get(url, timeout=25, max_bytes=3_000_000):
    req = urllib.request.Request(url, headers={
        "User-Agent": UA,
        "Accept": "text/html,application/xhtml+xml,application/pdf,*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read(max_bytes), r.headers.get("Content-Type", "?"), r.status


# ============================================================ vendor catalogs
# Qorvo's own navigation menu, read off the rendered page. These are the
# enumeration entry points -- one server-rendered table each.
QORVO_CATEGORIES = [
    "active-antenna-systems/beamformers",
    "active-antenna-systems/front-end-modules",
    "active-antenna-systems/if-transceivers",
    "amplifiers/distributed-amplifiers",
    "amplifiers/driver-amplifiers",
    "amplifiers/gain-block-amplifiers",
    "amplifiers/high-frequency-amplifiers",
    "amplifiers/low-noise-amplifiers",
    "amplifiers/low-phase-noise-amplifiers",
    "amplifiers/power-amplifiers",
    "amplifiers/spatium",
    "amplifiers/variable-gain-amplifiers",
    "control-products/attenuators",
    "control-products/limiters",
    "control-products/phase-shifters",
    "discrete-transistors/gaas-phemts",
    "discrete-transistors/gan-hemts",
    "filters-duplexers/rf-filters",
    "frequency-converters/mixers",
    "frequency-converters/modulators",
    "frequency-converters/multipliers",
    "frequency-converters/upconverters-downconverters",
    "frequency-converters/integrated-synthesizers-with-mixers",
    "integrated-products/defense-aerospace",
    "passives/fixed-attenuators",
    "passives/limiters",
    "passives/transformers",
    "switches/discrete-switches",
]

MARKI_FORMS = ["connectorized", "surface-mount", "bare-die", "waveguide"]
MARKI_CATEGORIES = ["mixers", "amplifiers", "filters", "couplers", "power-dividers",
                    "baluns", "bias-tees", "equalizers", "limiters", "multipliers",
                    "detectors", "attenuators"]

VENDOR_CATALOGS = OrderedDict([
    ("qorvo", {
        "name": "Qorvo",
        "status": "proven: server-rendered tables, PN + datasheet id per row",
        "urls": ["https://www.qorvo.com/products/product-list?categoryID=ca0118"]
                + [f"https://www.qorvo.com/products/{c}"
                   for c in QORVO_CATEGORIES[:4]],
        # v2 scanned only urls[0] for categoryIDs -- and urls[0] IS ca0118, so it
        # found only itself. The family landing pages carry ~305 links each and
        # are where the other ids live.
        "discover_seeds": [f"https://www.qorvo.com/products/{c}"
                           for c in ("amplifiers/low-noise-amplifiers",
                                     "amplifiers/power-amplifiers",
                                     "control-products/attenuators",
                                     "integrated-products/defense-aerospace",
                                     "frequency-converters/mixers")],
        "parser": "qorvo",
    }),
    ("marki", {
        "name": "Marki Microwave",
        "status": "researched: category pages list products; per-product datasheet page",
        "urls": [f"https://markimicrowave.com/products/{f}/mixers/"
                 for f in MARKI_FORMS[:2]]
                + ["https://markimicrowave.com/products/connectorized/amplifiers/"],
        "parser": "marki",
    }),
    ("adi", {
        "name": "Analog Devices",
        "status": ("NOT scraped: analog.com 403s and /en/products.html is a "
                   "23 kB shell. Part numbers come from your space xlsx instead."),
        "urls": [],
        "parser": "generic",
        "xlsx": True,
    }),
    ("macom", {
        "name": "MACOM",
        "status": "datasheet pattern known; LISTING page unproven -- probing candidates",
        "urls": ["https://www.macom.com/products",
                 "https://www.macom.com/products/rf-microwave"],
        "parser": "generic",
    }),
    ("skyworks", {
        "name": "Skyworks",
        "status": "datasheet pattern known; LISTING page unproven -- probing candidates",
        "urls": ["https://www.skyworksinc.com/en/Products",
                 "https://www.skyworksinc.com/en/Products/Amplifiers"],
        "parser": "generic",
    }),
])


# ================================================================== parsers
# All regex-based: no bs4 dependency, and a debug tool should not fail because
# of a missing package.

_HREF = re.compile(r'href\s*=\s*["\']([^"\']+)["\']', re.I)
_DOC_SUFFIX = re.compile(
    r"(?:[-_\s]+(?:data-?sheets?|ds|spec(?:ification)?s?|rev[a-z0-9.]*|final|"
    r"public|en|pdf)|_\d{6}[A-Z]?)$", re.I)


_WORDY = re.compile(r"^[A-Za-z]{4,}$")


def looks_like_pn(s):
    """Is this string a part number, or a subfamily slug?

    The v1 run pulled in DOUBLE-BALANCED-MIXERS, T3-HIGH-LINEARITY, LEGACY-MIXERS,
    TIMING-SERDES, Protections ... because any last path segment was accepted.
    Two rules separate them cleanly on every case that run produced:
      * every real part number contains a digit
      * no segment AFTER the first is an English word (4+ letters)
    The first segment is exempt because real prefixes look like words:
    MAAP-011325, MADT-011000, ADRF5545A.
    """
    s = (s or "").strip().strip("-_")
    if not (3 < len(s) <= 40):
        return False
    if not any(c.isdigit() for c in s):
        return False
    for seg in re.split(r"[-_]", s)[1:]:
        if _WORDY.match(seg):
            return False
    return True


def pn_from_filename(stem):
    """Part number out of a datasheet filename, keeping hyphenated segments.

    Splitting on the first hyphen truncated 'MM1-1140HS-GaAs Mixer' to 'MM1'.
    Vendor filenames append a description or document number, so trim only the
    segments that are clearly descriptive (they contain lower-case letters)."""
    s = urllib.parse.unquote(stem).strip()
    s = re.sub(r"\.pdf.*$", "", s, flags=re.I)
    s = s.split()[0] if s.split() else s
    for _ in range(3):
        s2 = _DOC_SUFFIX.sub("", s)
        if s2 == s:
            break
        s = s2
    if any(c.isupper() for c in s):
        kept = []
        for seg in re.split(r"([-_])", s):
            if seg in ("-", "_"):
                kept.append(seg)
                continue
            descriptive = (re.search(r"[a-z]{2}", seg)
                           or (any(c.islower() for c in seg)
                               and any(c.isupper() for c in seg)))
            if descriptive and kept:
                break
            kept.append(seg)
        s = "".join(kept).rstrip("-_")
    return s[:60]


def _abs(base, href):
    return urllib.parse.urljoin(base, href.strip())


def parse_qorvo(html, base):
    """Pair each /products/p/<PN> with the /products/d/<id> in the same row.

    The listing table renders the part link and (when one exists) the datasheet
    link adjacent to each other, so walking the document in order and attaching
    each d-link to the most recent p-link reconstructs the rows without needing
    to parse the table structure."""
    hits = []
    for m in re.finditer(
            r'href\s*=\s*["\']([^"\']*/products/(p|d)/([^"\'?#]+))["\']', html, re.I):
        hits.append((m.start(), m.group(2).lower(), m.group(3), m.group(1)))
    pairs, current = OrderedDict(), None
    for _pos, kind, ident, href in hits:
        if kind == "p":
            current = ident.strip()
            pairs.setdefault(current, {"pn": current, "product_url": _abs(base, href),
                                       "datasheet_url": None})
        elif kind == "d" and current:
            if pairs[current]["datasheet_url"] is None:
                pairs[current]["datasheet_url"] = _abs(base, href)
    return list(pairs.values())


def parse_marki(html, base):
    """Marki category pages link to /products/<form>/<cat>/<pn>/ and often to a
    /datasheet/ child or a direct /Assets/datasheets/<PN>.pdf."""
    out = OrderedDict()
    for href in _HREF.findall(html):
        h = href.strip()
        m = re.match(r"^(?:https?://[^/]+)?/products/([a-z-]+)/([a-z0-9-]+)/"
                     r"([A-Za-z0-9][A-Za-z0-9._-]*)/?$", h, re.I)
        if m and looks_like_pn(m.group(3)):
            pn = m.group(3).upper()
            out.setdefault(pn, {"pn": pn, "product_url": _abs(base, h),
                                "datasheet_url": None})
        if re.search(r"\.pdf($|[?#])", h, re.I):
            pn = pn_from_filename(h.rsplit("/", 1)[-1]).upper()
            if not looks_like_pn(pn):
                continue          # product catalogue PDFs, app notes, etc.
            rec = out.setdefault(pn, {"pn": pn, "product_url": None,
                                      "datasheet_url": None})
            rec["datasheet_url"] = rec["datasheet_url"] or _abs(base, h)
    # A product page with no PDF on the listing still has two known candidates:
    # its /datasheet/ child page, and the legacy flat Assets path.
    for rec in out.values():
        if not rec["datasheet_url"] and rec["product_url"]:
            rec["datasheet_url"] = rec["product_url"].rstrip("/") + "/datasheet/"
            rec["alt_datasheet_url"] = (
                "https://markimicrowave.com/Assets/datasheets/"
                + urllib.parse.quote(rec["pn"], safe="+-_.") + ".pdf")
    return list(out.values())


def parse_generic(html, base):
    """No vendor knowledge: every .pdf link, plus anything that looks like a
    product page carrying a part-number-ish last path segment. Enough to tell us
    whether a page is server-rendered and worth writing a real parser for."""
    out = OrderedDict()
    for href in _HREF.findall(html):
        h = href.strip()
        if re.search(r"\.pdf($|[?#])", h, re.I):
            pn = pn_from_filename(h.rsplit("/", 1)[-1])
            if looks_like_pn(pn):
                out.setdefault(pn.upper(), {"pn": pn, "product_url": None,
                                            "datasheet_url": _abs(base, h)})
    # v2 only matched a PN sitting immediately after /products/, so it missed
    # both real shapes found by research:
    #   MACOM     /products/product-detail/MAAP-011325
    #   Skyworks  /en/Products/Amplifiers/SKY66318-11
    # Take the LAST path segment of any same-host product URL and test that.
    host = urllib.parse.urlparse(base).netloc
    for href in _HREF.findall(html):
        u = _abs(base, href.strip())
        pr = urllib.parse.urlparse(u)
        if pr.netloc != host or not re.search(r"/products?/", pr.path, re.I):
            continue
        if re.search(r"\.(pdf|jpg|png|zip|xlsx?)($|\?)", pr.path, re.I):
            continue
        last = urllib.parse.unquote(pr.path.rstrip("/").rsplit("/", 1)[-1])
        if looks_like_pn(last):
            pn = last.upper()
            out.setdefault(pn, {"pn": pn, "product_url": u,
                                "datasheet_url": None})
    return list(out.values())


PARSERS = {"qorvo": parse_qorvo, "marki": parse_marki, "generic": parse_generic}


def derive_datasheet_urls(vendor_key, pn):
    """Datasheet URLs built from the patterns confirmed earlier by fetching real
    files. For MACOM and ADI the pattern IS the answer -- no page visit needed.
    Skyworks Si* parts follow a pattern; SKY* carry an opaque document number, so
    their product page has to be followed."""
    n = pn.strip()
    low, up = n.lower(), n.upper()
    q = urllib.parse.quote(n, safe="+-_.")
    if vendor_key == "macom":
        return [f"https://cdn.macom.com/datasheets/{q}.pdf"]
    if vendor_key == "adi":
        base = "https://www.analog.com/media/en/technical-documentation/data-sheets/"
        cands = [f"{base}{low}.pdf", f"{base}{up}.pdf"]
        # 'AD1671S' is the space variant of 'AD1671'; the datasheet is often
        # filed under the base part number.
        stem = re.sub(r"S$", "", n, flags=re.I)
        if stem and stem.lower() != low:
            cands += [f"{base}{stem.lower()}.pdf", f"{base}{stem.upper()}.pdf"]
        return cands
    if vendor_key == "skyworks":
        base = ("https://www.skyworksinc.com/-/media/Skyworks/SL/documents/"
                "public/data-sheets/")
        return [f"{base}{low}-datasheet.pdf", f"{base}{low}.pdf"]
    if vendor_key == "marki":
        return [f"https://markimicrowave.com/Assets/datasheets/{q}.pdf"]
    if vendor_key == "minicircuits":
        return [f"https://www.minicircuits.com/pdfs/{q}.pdf"]
    return []


def looks_js_shell(html):
    """A JS-only page: little text, few links, and a framework root element."""
    links = len(_HREF.findall(html))
    text = re.sub(r"<script.*?</script>", " ", html, flags=re.S | re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    words = len(text.split())
    markers = any(m in html.lower() for m in
                  ('id="root"', 'id="app"', "__next_data__", "ng-app",
                   "data-reactroot"))
    # Heuristic, and only meaningful on a real page: a small test fixture is
    # legitimately short. Require some page weight before calling it a JS shell.
    substantial = len(html) > 5000
    js = substantial and ((links < 20 and words < 300)
                          or (markers and links < 30 and words < 600))
    return js, links, words


# ================================================================ pdf stage
def pdf_text(blob, pages=2):
    try:
        import warnings
        import pdfplumber
        warnings.filterwarnings("ignore")
        with pdfplumber.open(io.BytesIO(blob)) as pdf:
            return "\n".join((p.extract_text() or "") for p in pdf.pages[:pages])
    except Exception as e:
        try:
            from pypdf import PdfReader
            r = PdfReader(io.BytesIO(blob))
            return "\n".join((p.extract_text() or "") for p in r.pages[:pages])
        except Exception:
            return f"__EXTRACT_FAILED__ {type(e).__name__}"


# ==================================================== discovery + following
# The v1 run showed MACOM /products and Skyworks /en/Products return only
# CATEGORY links, and Qorvo's parametric tables live solely behind
# product-list?categoryID=<id> -- the /products/<family>/<sub> pages are landing
# pages with no table. So the missing capability is: find the listing URLs from a
# seed page, then follow them.

_QORVO_CATID = re.compile(r"categoryID=([A-Za-z0-9]+)", re.I)


def discover_listing_urls(vendor_key, html, base, limit=6):
    """Candidate listing/table URLs discovered from a seed page."""
    host = urllib.parse.urlparse(base).netloc
    found = []

    def add(u):
        if u not in found and urllib.parse.urlparse(u).netloc == host:
            found.append(u)

    if vendor_key == "qorvo":
        # the only shape that carries a parametric table
        for cid in dict.fromkeys(_QORVO_CATID.findall(html)):
            add(f"https://www.qorvo.com/products/product-list?categoryID={cid}")
        return found[:limit]

    # generic: same-host links that look like a product category, preferring
    # deeper paths (a subcategory is likelier to hold the actual table)
    cands = []
    for href in _HREF.findall(html):
        u = _abs(base, href.strip())
        pr = urllib.parse.urlparse(u)
        if pr.netloc != host or not re.search(r"/products?/", pr.path, re.I):
            continue
        if re.search(r"\.(pdf|jpg|png|zip|xlsx?)($|\?)", pr.path, re.I):
            continue
        last = pr.path.rstrip("/").rsplit("/", 1)[-1]
        if looks_like_pn(last):
            continue                      # that is a part page, not a category
        cands.append((pr.path.count("/"), u))
    for _depth, u in sorted(cands, key=lambda t: -t[0]):
        add(u)
    return found[:limit]


_DATEISH = re.compile(r"^\d{4}-\d{2}-\d{2}([ T]|$)|^\d{1,2}/\d{1,2}/\d{2,4}$")
_CATALOGUE = re.compile(r"catalog|catalogue|brochure|selection|short.?form|"
                        r"price|guide|portfolio", re.I)


def pdf_candidate_score(url, pn):
    """Rank PDF links found on a landing page.

    v2 took the FIRST .pdf link, which on Marki's /datasheet/ page is
    MM_Catalog_Connectorized_Waveguide_7-2026.pdf -- the 2.9 MB full catalogue,
    returned identically for MM1-0320LBH and MM1-1886HM. A part's own datasheet
    almost always carries its part number in the filename, and a catalogue never
    does."""
    name = urllib.parse.unquote(url.rsplit("/", 1)[-1])
    flat = re.sub(r"[^A-Z0-9]", "", name.upper())
    key = re.sub(r"[^A-Z0-9]", "", (pn or "").upper())
    score = 0
    if key and key in flat:
        score += 100
    elif key and len(key) > 5 and key[:6] in flat:
        score += 40                      # same family, e.g. a shared datasheet
    if _CATALOGUE.search(name):
        score -= 200
    return score


def resolve_pdf(url, pn=None, depth=0):
    """Fetch a datasheet URL, following an HTML landing page to the real PDF.

    Qorvo's /products/d/<id> served text/html for da006470 but a real %PDF for
    da009265 -- the same URL shape does both, so the HTML case has to be followed
    rather than written off as a failure."""
    blob, ctype, code = http_get(url)
    if blob[:5] == b"%PDF-":
        return blob, url, f"HTTP {code} direct"
    if depth >= 1:
        return None, url, f"HTTP {code} {ctype.split(';')[0]} (not a PDF)"
    html = blob.decode("utf-8", "replace")
    pdfs = [_abs(url, h) for h in _HREF.findall(html)
            if re.search(r"\.pdf($|[?#])", h, re.I)]
    # some sites embed the file rather than link it
    pdfs += [_abs(url, m.group(1)) for m in
             re.finditer(r'(?:src|data)\s*=\s*["\']([^"\']+\.pdf[^"\']*)["\']',
                         html, re.I)]
    ranked = sorted(dict.fromkeys(pdfs),
                    key=lambda u: -pdf_candidate_score(u, pn))
    for cand in ranked:
        sc = pdf_candidate_score(cand, pn)
        if sc <= 0:
            # Nothing on the page names this part. Refuse rather than hand back
            # a catalogue that would poison the training data for every part.
            break
        try:
            blob2, ctype2, code2 = http_get(cand)
        except Exception:
            continue
        if blob2[:5] == b"%PDF-":
            return blob2, cand, f"followed HTML -> {cand.rsplit('/', 1)[-1]}"
    if ranked:
        best = ranked[0].rsplit("/", 1)[-1]
        return None, url, (f"HTML had {len(ranked)} PDF link(s) but none name this "
                           f"part (best was '{best[:48]}') -- refusing a catalogue")
    snippet = " ".join(re.sub(r"<[^>]+>", " ", html).split())[:110]
    return None, url, f"HTML with no PDF link; page says: \"{snippet}\""


def stage_discover(vendors, follow):
    say("\n" + "=" * 76)
    say(f"DISCOVERY: find listing URLs from a seed page, then follow up to "
        f"{follow}")
    say("=" * 76)
    say("  This is the capability the first run showed was missing: MACOM and")
    say("  Skyworks seed pages list only categories, and Qorvo's tables live")
    say("  only behind product-list?categoryID=<id>.\n")
    extra = {}
    for key, cfg in vendors.items():
        seeds = cfg.get("discover_seeds") or cfg.get("urls") or []
        if not seeds:
            say(f"  {cfg['name']:<18} no seed pages (by design)")
            continue
        urls, html_any = [], None
        for seed in seeds:
            status, value, secs = guarded(f"seed-{key}",
                                          lambda u=seed: http_get(u), 35)
            if status != "ok":
                say(f"  {cfg['name']:<18} seed failed ({seed.rsplit('/', 1)[-1]}): "
                    f"{value}")
                continue
            html_any = value[0].decode("utf-8", "replace")
            for u in discover_listing_urls(key, html_any, seed, limit=follow * 3):
                if u not in urls:
                    urls.append(u)
            if len(urls) >= follow:
                break
        urls = urls[:follow]
        if not urls:
            continue
        say(f"  {cfg['name']:<18} discovered {len(urls)} candidate listing URL(s) "
            f"from {len(seeds)} seed page(s)")
        for u in urls:
            say(f"      {u}")
        for u in urls:
            st, val, sec = guarded(f"follow-{key}", lambda x=u: http_get(x), 35)
            if st != "ok":
                say(f"    [{st:<7}] {sec:5.1f}s  {u}")
                say(f"               -> {val}")
                continue
            h2 = val[0].decode("utf-8", "replace")
            recs = PARSERS[cfg["parser"]](h2, u)
            recs = [r for r in recs if looks_like_pn(r["pn"])]
            with_ds = sum(1 for r in recs if r.get("datasheet_url"))
            js, links, words = looks_js_shell(h2)
            say(f"    [ok     ] {sec:5.1f}s  {len(val[0]) // 1024} kB  "
                f"{links} links  {words} words  -> {len(recs)} part(s), "
                f"{with_ds} with datasheet"
                + ("   <-- looks JS-rendered" if js else ""))
            for r in recs[:6]:
                say(f"                 {r['pn']:<24} "
                    f"{(r.get('datasheet_url') or '(none)')[:66]}")
            if recs:
                extra.setdefault(key, []).extend(recs)
    return extra


# =================================================================== stages
def stage_catalog(vendors, local_dir):
    say("=" * 76)
    say("CATALOG PAGES: fetch -> render check -> parse")
    say("=" * 76)
    found = {}
    if local_dir:
        say(f"  --local: parsing HTML files in {local_dir} (no requests made)\n")
        files = sorted(Path(local_dir).rglob("*.htm*"))
        if not files:
            say("  no .html/.htm files there.")
            return found
        for fp in files:
            html = fp.read_text(encoding="utf-8", errors="replace")
            guess = ("qorvo" if "qorvo.com" in html else
                     "marki" if "markimicrowave" in html else "generic")
            recs = PARSERS[guess](html, "https://example.invalid/")
            with_ds = sum(1 for r in recs if r.get("datasheet_url"))
            say(f"  {fp.name:<38} parser={guess:<8} {len(recs):>4} part(s), "
                f"{with_ds:>4} with a datasheet link")
            for r in recs[:6]:
                say(f"      {r['pn']:<22} {r.get('datasheet_url') or '(no ds link)'}")
            found.setdefault(guess, []).extend(recs)
        return found

    for key, cfg in vendors.items():
        say(f"\n  {cfg['name']}  [{key}]")
        say(f"    {cfg['status']}")
        for url in cfg["urls"]:
            status, value, secs = guarded(f"cat-{key}", lambda u=url: http_get(u), 35)
            if status != "ok":
                say(f"    [{status:<7}] {secs:5.1f}s  {url}")
                say(f"               -> {value}")
                continue
            blob, ctype, code = value
            html = blob.decode("utf-8", "replace")
            js, links, words = looks_js_shell(html)
            recs = PARSERS[cfg["parser"]](html, url)
            with_ds = sum(1 for r in recs if r.get("datasheet_url"))
            say(f"    [ok     ] {secs:5.1f}s  HTTP {code}  {len(blob) // 1024} kB  "
                f"{links} links, {words} words"
                + ("   <-- looks JS-rendered" if js else ""))
            say(f"               {url}")
            say(f"               parsed {len(recs)} part(s), {with_ds} with a "
                f"datasheet link")
            for r in recs[:8]:
                say(f"                 {r['pn']:<24} "
                    f"{(r.get('datasheet_url') or '(none)')[:70]}")
            if recs:
                found.setdefault(key, []).extend(recs)
    return found


def dedupe(recs):
    """One record per part. The four Marki subfamily URLs each returned the same
    page, so the same 97 parts were counted four times (593 total)."""
    seen, out = set(), []
    for r in recs:
        k = re.sub(r"[^A-Z0-9]", "", (r.get("pn") or "").upper())
        if not k or k in seen:
            continue
        seen.add(k)
        out.append(r)
    return out


def stage_adi_xlsx(path, n_pdf):
    """ADI without scraping: read part numbers from the space portfolio sheet and
    build datasheet URLs from the confirmed pattern."""
    say("\n" + "=" * 76)
    say("ADI: part numbers from the space portfolio spreadsheet (no scraping)")
    say("=" * 76)
    p = Path(path)
    if not p.is_file():
        say(f"  not found: {p}")
        say("  pass --adi-xlsx <path> if it lives elsewhere")
        return []
    try:
        import openpyxl
    except ImportError:
        say("  openpyxl not installed:  pip install openpyxl")
        return []
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[(c.value if c.value is not None else "") for c in r]
            for r in ws.iter_rows(max_row=400)]
    wb.close()
    # Pick the column by HEADER NAME. The heuristic picked column 2, "Orderable
    # Material Number" (AD1671F-EMX, 5962R8853901V2A) -- those are order codes and
    # DLA drawing numbers with no datasheet of their own. Column 0 is "Generic
    # Part Number" (AD1671S), which is what an analog.com datasheet URL uses.
    header_row, best_col = None, None
    for ri, r in enumerate(rows[:20]):
        for ci, cell in enumerate(r):
            if re.sub(r"\s+", " ", str(cell)).strip().lower() == "generic part number":
                header_row, best_col = ri, ci
                break
        if best_col is not None:
            break
    label = "Generic Part Number"
    if best_col is None:
        label = "(heuristic: most part-number-like column)"
        best_hits = 0
        for ci in range(max((len(r) for r in rows), default=0)):
            hits = sum(1 for r in rows[1:] if ci < len(r)
                       and looks_like_pn(str(r[ci])))
            if hits > best_hits:
                best_hits, best_col, header_row = hits, ci, 0
        if best_col is None or best_hits < 5:
            say(f"  could not find a part-number column")
            return []
    pns, seen = [], set()
    for r in rows[(header_row or 0) + 1:]:
        if best_col >= len(r):
            continue
        v = str(r[best_col]).strip()
        if _DATEISH.match(v):
            continue                      # 'Last Modified Date' cells etc.
        if looks_like_pn(v) and v.upper() not in seen:
            seen.add(v.upper())
            pns.append(v)
    say(f"  {p.name}: sheet '{ws.title}', column {best_col} = {label}")
    say(f"  {len(pns)} distinct part number(s); first 10: "
        f"{', '.join(pns[:10])}")
    recs = [{"pn": n, "product_url": None,
             "datasheet_url": derive_datasheet_urls("adi", n)[0],
             "from_pattern": True} for n in pns]
    return recs


def stage_pdf(found, n_per_vendor):
    say("\n" + "=" * 76)
    say("DATASHEET PDFs: fetch -> %PDF magic -> text extraction")
    say("=" * 76)
    if not found:
        say("  nothing to try (no datasheet URLs were parsed).")
        return
    for key, recs in found.items():
        recs = dedupe(recs)
        # fall back to the confirmed URL pattern when the page gave no link
        for r in recs:
            if not r.get("datasheet_url"):
                cand = derive_datasheet_urls(key, r["pn"])
                if cand:
                    r["datasheet_url"] = cand[0]
                    r["candidates"] = cand
                    r["from_pattern"] = True
        with_ds = [r for r in recs if r.get("datasheet_url")]
        if not with_ds:
            say(f"\n  {key}: no datasheet URLs parsed, nothing to fetch")
            continue
        say(f"\n  {key}: trying {min(n_per_vendor, len(with_ds))} of "
            f"{len(with_ds)} datasheet URL(s)")
        for r in with_ds[:n_per_vendor]:
            cands = r.get("candidates") or [r["datasheet_url"]]

            def try_all(urls=tuple(cands), n=r["pn"]):
                last = None
                for u in urls:
                    try:
                        blob, final, how = resolve_pdf(u, n)
                    except Exception as e:
                        last = f"{type(e).__name__} on {u.rsplit('/', 1)[-1]}"
                        continue
                    if blob:
                        return blob, final, how
                    last = how
                return None, urls[0], (f"{len(urls)} candidate URL(s) tried; "
                                       f"last: {last}")
            status, value, secs = guarded(f"pdf-{r['pn']}", try_all, 90)
            if status != "ok":
                say(f"    [{status:<7}] {secs:5.1f}s  {r['pn']}  -> {value}")
                continue
            blob, final_url, how = value
            if not blob:
                say(f"    [no-pdf ] {secs:5.1f}s  {r['pn']:<20} {how}")
                continue
            say(f"    [ok     ] {secs:5.1f}s  {r['pn']:<20} "
                f"{len(blob) // 1024} kB  %PDF  ({how})")
            txt = pdf_text(blob)
            if txt.startswith("__EXTRACT_FAILED__"):
                say(f"               text extraction FAILED: {txt}")
                continue
            words = len(txt.split())
            say(f"               extracted {len(txt)} chars / {words} words")
            snippet = " ".join(txt.split())[:150]
            say(f"               \"{snippet}\"")
            # does it contain the construction language the classifier needs?
            cues = [w for w in ("hermetic", "ceramic", "GaAs", "GaN", "alumina",
                                "bare die", "MIL-", "screened", "laser weld",
                                "plastic", "epoxy", "-55")
                    if re.search(re.escape(w), txt, re.I)]
            say(f"               construction cues present: "
                f"{', '.join(cues) if cues else 'NONE'}")


def stage_minicircuits(catalog_path, n):
    """Control case: we already hold this catalog, so it proves the PDF chain."""
    say("\n" + "=" * 76)
    say("CONTROL: Mini-Circuits (local JSON catalog -> PDF)")
    say("=" * 76)
    p = Path(catalog_path) if catalog_path else None
    if not p or not p.is_file():
        say(f"  catalog not found ({catalog_path}); skipping the control case")
        return
    import json
    data = json.loads(p.read_text(encoding="utf-8"))
    if isinstance(data, dict):
        data = next(v for v in data.values() if isinstance(v, list))
    say(f"  loaded {len(data)} parts from {p.name}")
    picks = [r for r in data if r.get("datasheet_url")][:n]
    for r in picks:
        url = r["datasheet_url"]
        status, value, secs = guarded(f"mc-{r['pn']}", lambda u=url: http_get(u), 40)
        if status != "ok":
            say(f"    [{status:<7}] {secs:5.1f}s  {r['pn']}  -> {value}")
            continue
        blob, ctype, code = value
        ok = blob[:5] == b"%PDF-"
        txt = pdf_text(blob) if ok else ""
        say(f"    [ok     ] {secs:5.1f}s  {r['pn']:<18} {len(blob) // 1024} kB  "
            f"{'%PDF' if ok else 'NOT PDF'}  "
            f"{len(txt.split()) if txt else 0} words extracted")


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--report", default="spacequal_catalog_report.txt")
    ap.add_argument("--vendor", help="only this vendor key")
    ap.add_argument("--local", help="folder of catalog pages you saved yourself")
    ap.add_argument("--pdfs", type=int, default=2,
                    help="datasheet PDFs to try per vendor (0 = none)")
    ap.add_argument("--no-pdf", action="store_true")
    ap.add_argument("--adi-xlsx",
                    default=str(Path.home() / "Downloads" / "newSources"
                               / "adi_space_portfolio_2026-07-28.xlsx"),
                    help="ADI space portfolio spreadsheet (part-number source)")
    ap.add_argument("--follow", type=int, default=4,
                    help="listing URLs to discover and follow per vendor (0 = off)")
    ap.add_argument("--catalog", help="Mini-Circuits products JSON (control case)",
                    default=str(Path.home() / "Downloads" / "rfparts" / "rfparts"
                               / "sources" / "minicircuits_products_full.json"))
    args = ap.parse_args()

    say("spacequal catalog + datasheet parsing debug")
    say(f"generated {time.strftime('%Y-%m-%d %H:%M:%S')}")
    say(f"python {sys.version.split()[0]} on {sys.platform}")
    try:
        import pdfplumber
        say(f"pdfplumber {pdfplumber.__version__}")
    except ImportError:
        say("pdfplumber MISSING -- PDF text extraction will fall back to pypdf")
    say("")

    vendors = VENDOR_CATALOGS
    if args.vendor:
        if args.vendor not in vendors:
            raise SystemExit(f"unknown vendor '{args.vendor}'. "
                             f"choose from: {', '.join(vendors)}")
        vendors = OrderedDict([(args.vendor, vendors[args.vendor])])

    found = stage_catalog(vendors, args.local)
    if args.follow and not args.local:
        extra = stage_discover(vendors, args.follow)
        for k, v in extra.items():
            found.setdefault(k, []).extend(v)
    if not args.local and "adi" in vendors:
        adi = stage_adi_xlsx(args.adi_xlsx, args.pdfs)
        if adi:
            found["adi"] = adi
    if not args.no_pdf and args.pdfs:
        stage_pdf(found, args.pdfs)
        if not args.local:
            stage_minicircuits(args.catalog, 2)

    say("\n" + "=" * 76)
    say("SCORECARD")
    say("=" * 76)
    say(f"  {'vendor':<18} {'parts':>7} {'with ds':>8}   verdict")
    for key in vendors:
        recs = found.get(key, [])
        real = dedupe([r for r in recs if looks_like_pn(r["pn"])])
        with_ds = sum(1 for r in real if r.get("datasheet_url"))
        if with_ds:
            verdict = "READY to wire in"
        elif real:
            verdict = "parts found, no datasheet links -- needs another hop"
        else:
            verdict = "no parts parsed -- save pages and use --local"
        say(f"  {VENDOR_CATALOGS[key]['name']:<18} {len(real):>7} {with_ds:>8}   "
            f"{verdict}")
    say("")
    say("WHAT TO CONCLUDE")
    say("=" * 76)
    say("  For each vendor, three things had to work:")
    say("    fetch     the listing page came back at all")
    say("    parse     part numbers AND datasheet links came out of it")
    say("    pdf       one of those links returned a real %PDF with readable text")
    say("  A vendor is only worth wiring in when all three pass. If a page came")
    say("  back but parsed 0 parts and says 'looks JS-rendered', its catalog is")
    say("  built in the browser and the answer is to save pages from your own")
    say("  Chrome and re-run with --local.")

    out = Path(args.report)
    try:
        out.write_text(OUT.getvalue(), encoding="utf-8")
        print(f"\nreport written to {out.resolve()}", flush=True)
    except Exception as e:
        print(f"\ncould not write {out}: {e}", flush=True)
    sys.stdout.flush()
    os._exit(0)


if __name__ == "__main__":
    main()
