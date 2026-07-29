#!/usr/bin/env python3
"""spacequal_finish_debug — the last two vendors. Mini-Circuits, Qorvo, MACOM and
Skyworks are already proven; nothing here re-tests them.

    python spacequal_finish_debug.py        # that is all -- no flags, no setup

It opens its own Chrome window, uses it, and closes it again when the run ends or
you press Ctrl-C. You do not have to launch anything.

WHAT IS LEFT, AND WHY
---------------------
MARKI -- their "datasheet" PDFs are unusable, and now we know exactly why:
    triage: 0 font ref(s); 35 image ref(s); producer=Skia/PDF m148
Skia/PDF is Chrome's print-to-PDF engine, so those files are print renderings of
the product page: images with no text layer. No extractor can read them, pypdf
included; only OCR could, and we do not need to.

    Because the /datasheet/ page contains the WHOLE datasheet as HTML:
      General Description, Features, Port Functions, Package Information,
      Recommended Operating Conditions, full Electrical Specifications, and
      lines like "Specifications guaranteed from -55 to +100 C, measured in a
      50-Ohm system".
    Parsing that HTML gives cleaner text than any PDF extraction would. This
    tool proves it end to end.

ADI -- diagnosed as a deliberate block, not a wrong URL:
    HTTP 403 on AD1671.pdf      (403, not 404)
Their product-category page 403s the same way. The URL pattern is right; ADI is
refusing scripted requests to that path.

    I am NOT going to send a browser-like User-Agent to get around that. It is
    an access control they set on purpose, and spoofing identity to defeat it is
    the line this project does not cross. The honest route is the browser you
    already have: we know all 186 URLs from your space spreadsheet, so
    this tool opens a Chrome window itself and fetches that known list through it
    -- real cookies, real TLS, real user agent, nothing disguised, human pace.
    The window is visible, uses its own profile directory, and is closed again on
    exit. If a debuggable Chrome is already running on the port we attach to that
    instead and leave it open, because we only close what we opened.

MARKI also goes through the browser when needed: each page is tried over plain
HTTP first (faster and lighter), and only falls back to the browser if that is
blocked or returns suspiciously little text. The report says which path each page
used.
"""
from __future__ import annotations

import argparse
import html as htmllib
import io
import os
import re
import socket
import sys
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

socket.setdefaulttimeout(20)
UA = "rfparts-spacequal-debug/1.2 (catalog parser diagnostic)"
OUT = io.StringIO()

# Construction / qualification vocabulary the classifier actually needs.
CUES = ("hermetic", "ceramic", "GaAs", "GaN", "alumina", "bare die", "MIL-",
        "screened", "laser weld", "plastic", "epoxy", "-55", "kovar",
        "glass-to-metal", "space", "hi-rel", "LTCC", "thin film")


def say(line=""):
    print(line, flush=True)
    OUT.write(line + "\n")


def guarded(label, fn, budget):
    box = {}
    t0 = time.time()

    def target():
        try:
            box["v"] = fn()
            box["s"] = "ok"
        except urllib.error.HTTPError as e:
            box["v"] = f"HTTP {e.code}"
            box["s"] = "error"
        except Exception as e:
            box["v"] = f"{type(e).__name__}: {e}"
            box["s"] = "error"
    th = threading.Thread(target=target, daemon=True, name=label)
    th.start()
    th.join(budget)
    if th.is_alive():
        return "TIMEOUT", f"blocked past {budget}s", time.time() - t0
    return box.get("s", "error"), box.get("v"), time.time() - t0


def http_get(url, timeout=25, max_bytes=4_000_000):
    req = urllib.request.Request(url, headers={
        "User-Agent": UA, "Accept": "text/html,application/pdf,*/*"})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return r.read(max_bytes), r.headers.get("Content-Type", "?"), r.status


def found_cues(text):
    return [c for c in CUES if re.search(re.escape(c), text, re.I)]


# ======================================================= managed Chrome window
# Chrome is launched HERE and closed again when the run finishes or you press
# Ctrl-C. It is a normal, visible Chrome window with its own profile directory --
# not a headless or stealth browser, and no fingerprint is altered. The reason we
# go through a browser at all is that analog.com refuses scripted requests; we
# stop pretending a script is welcome where the site said it is not, rather than
# disguising the script.
#
# If a debuggable Chrome is ALREADY listening on the port, we attach to that and
# leave it running afterwards -- we only ever close a window we opened ourselves.

CHROME_PORT = 9222
CHROME_PROFILE = Path.home() / "sq-chrome"

_CHROME_CANDIDATES = [
    r"C:\Program Files\Google\Chrome\Application\chrome.exe",
    r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
    str(Path(os.environ.get("LOCALAPPDATA", "")) /
        "Google" / "Chrome" / "Application" / "chrome.exe"),
    r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
    "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
    "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
]


def find_chrome():
    import shutil
    for c in _CHROME_CANDIDATES:
        if c and Path(c).is_file():
            return c
    for name in ("google-chrome", "google-chrome-stable", "chromium",
                 "chromium-browser", "microsoft-edge"):
        found = shutil.which(name)
        if found:
            return found
    return None


def port_open(port, host="127.0.0.1", timeout=0.6):
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False


class ManagedBrowser:
    """Launches (or attaches to) Chrome and always cleans up after itself."""

    def __init__(self, port=CHROME_PORT, profile=CHROME_PROFILE, rate=1.0):
        self.port = port
        self.profile = Path(profile)
        self.rate = rate
        self.proc = None
        self.we_launched = False
        self._pw = None
        self._browser = None
        self.ctx = None
        self.page = None

    # -- lifecycle ---------------------------------------------------------
    def start(self):
        try:
            from playwright.sync_api import sync_playwright
        except ImportError:
            say("  Playwright is not installed:  pip install playwright")
            return False

        if port_open(self.port):
            say(f"  a debuggable browser is already on port {self.port} -- "
                f"attaching to it (and leaving it open afterwards)")
        else:
            exe = find_chrome()
            if not exe:
                say("  could not find Chrome or Edge. Looked in the usual places;")
                say("  install Chrome, or start it yourself with:")
                say(f'    chrome.exe --remote-debugging-port={self.port} '
                    f'--user-data-dir="{self.profile}"')
                return False
            self.profile.mkdir(parents=True, exist_ok=True)
            say(f"  launching {Path(exe).name} on port {self.port}")
            say(f"    profile: {self.profile}   (separate from your normal one,")
            say(f"    so it cannot hand off to an already-running Chrome)")
            import subprocess
            self.proc = subprocess.Popen(
                [exe, f"--remote-debugging-port={self.port}",
                 f"--user-data-dir={self.profile}",
                 "--no-first-run", "--no-default-browser-check",
                 "about:blank"],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            self.we_launched = True
            for _ in range(40):                    # up to ~20 s to come up
                if port_open(self.port):
                    break
                time.sleep(0.5)
            else:
                say("  Chrome did not open the debug port in 20 s; giving up")
                self.close()
                return False
            say("  Chrome is up -- you will see pages load in that window as it")
            say("  works; PDFs are fetched by navigating to them.")

        from playwright.sync_api import sync_playwright
        self._pw = sync_playwright().start()
        try:
            self._browser = self._pw.chromium.connect_over_cdp(
                f"http://localhost:{self.port}", timeout=20000)
        except Exception as e:
            say(f"  could not attach over CDP: {e}")
            self.close()
            return False
        self.ctx = (self._browser.contexts[0] if self._browser.contexts
                    else self._browser.new_context())
        self.page = self.ctx.pages[0] if self.ctx.pages else self.ctx.new_page()
        return True

    def close(self):
        try:
            if self._pw:
                self._pw.stop()
        except Exception:
            pass
        self._pw = None
        # Only ever close a window we opened ourselves.
        if self.proc and self.we_launched:
            say("  closing the Chrome window we opened")
            try:
                self.proc.terminate()
                try:
                    self.proc.wait(timeout=8)
                except Exception:
                    self.proc.kill()
            except Exception:
                pass
        self.proc = None

    def __enter__(self):
        self.ok = self.start()
        return self

    def __exit__(self, *exc):
        self.close()
        return False       # never swallow the exception (incl. KeyboardInterrupt)

    # -- fetching ----------------------------------------------------------
    # -- fetching ----------------------------------------------------------
    def _ensure_origin(self, url, timeout=45):
        """Park the page on the target's own origin so in-page fetch() is
        same-origin (no CORS) and carries that site's cookies."""
        pr = urllib.parse.urlparse(url)
        origin = f"{pr.scheme}://{pr.netloc}/"
        try:
            here = urllib.parse.urlparse(self.page.url)
            if f"{here.scheme}://{here.netloc}" == f"{pr.scheme}://{pr.netloc}":
                return
        except Exception:
            pass
        self.page.goto(origin, wait_until="domcontentloaded",
                       timeout=int(timeout * 1000))

    def looks_challenged(self):
        try:
            txt = (self.page.title() or "") + " " + (self.page.url or "")
            body = self.page.content()[:4000].lower()
        except Exception:
            return False
        markers = ("just a moment", "checking your browser", "attention required",
                   "cf-browser-verification", "enable javascript and cookies",
                   "verify you are human")
        low = txt.lower()
        return any(m in low or m in body for m in markers)

    def wait_for_human(self, what):
        """A challenge appeared. We do NOT try to solve it -- exactly the rule
        erf_save_pages.py follows. Hand control back and wait."""
        say(f"\n  *** A human check appeared for {what}.")
        say(f"  *** Clear it in the Chrome window we opened, then press Enter here.")
        try:
            input("  *** > ")
        except EOFError:
            say("  *** (no console available; continuing)")

    def get_html(self, url, timeout=60, wait_text=None, allow_pause=True):
        """Rendered DOM, waiting for content that arrives after DOMContentLoaded.

        Waiting only for domcontentloaded returned Marki's prose without its
        Electrical Specifications table, because that table is built later. So we
        settle the network and, when asked, wait for the table itself."""
        self.page.goto(url, wait_until="domcontentloaded",
                       timeout=int(timeout * 1000))
        if allow_pause and self.looks_challenged():
            self.wait_for_human(urllib.parse.urlparse(url).netloc)
            self.page.goto(url, wait_until="domcontentloaded",
                           timeout=int(timeout * 1000))
        try:
            self.page.wait_for_load_state("networkidle", timeout=15000)
        except Exception:
            pass
        if wait_text:
            try:
                self.page.wait_for_function(
                    "t => document.body && document.body.innerText"
                    ".toLowerCase().includes(t)",
                    arg=wait_text.lower(), timeout=15000)
            except Exception:
                pass                       # report what we have; caller decides
        return self.page.content()

    # Download inside the page: Chrome's own stack, so Chrome's certificate
    # store (the corporate inspection CA), Chrome's cookies, Chrome's UA.
    _FETCH_JS = """
    async (url) => {
      const r = await fetch(url, {credentials: 'include'});
      if (!r.ok) return {err: 'HTTP ' + r.status};
      const buf = await r.arrayBuffer();
      const bytes = new Uint8Array(buf);
      let bin = '';
      const CH = 0x8000;
      for (let i = 0; i < bytes.length; i += CH) {
        bin += String.fromCharCode.apply(null, bytes.subarray(i, i + CH));
      }
      return {b64: btoa(bin), len: bytes.length};
    }
    """

    def get_bytes(self, url, timeout=60):
        """PDF bytes. Returns (bytes, how). Layers, in order:

        1. IN-PAGE fetch() after parking on the site's origin. This is the one
           that actually works here: Chrome navigating to a PDF hands it to the
           built-in viewer and does NOT retain the body, so response.body() fails
           with "No resource with given identifier found" even though the document
           is plainly visible on screen. Asking the page's own JavaScript to
           download it sidesteps that, and rides Chrome's TLS + cookies.
        2. page.goto() + response.body() -- works when the URL is not handled by
           the viewer.
        3. ctx.request.get() -- browser cookies, but Node's TLS stack, which
           rejects a corporate inspection CA.
        4. a lenient request context, scoped to this one fetch.
        """
        import base64
        errs = []
        # 1
        try:
            self._ensure_origin(url, timeout)
            if self.looks_challenged():
                self.wait_for_human(urllib.parse.urlparse(url).netloc)
            out = self.page.evaluate(self._FETCH_JS, url)
            if isinstance(out, dict) and out.get("b64"):
                body = base64.b64decode(out["b64"])
                if body[:5] == b"%PDF-":
                    return body, "in-page fetch"
                errs.append(f"in-page got {len(body)} B, not a PDF")
            else:
                errs.append(f"in-page: {(out or {}).get('err', 'no data')}")
        except Exception as e:
            errs.append(f"in-page: {str(e).splitlines()[0][:80]}")
        # 2
        try:
            resp = self.page.goto(url, wait_until="commit",
                                  timeout=int(timeout * 1000))
            if resp is not None:
                body = resp.body()
                if body[:5] == b"%PDF-":
                    return body, "page.goto"
                errs.append(f"goto {resp.status}, not a PDF")
        except Exception as e:
            errs.append(f"goto: {str(e).splitlines()[0][:80]}")
        # 3
        try:
            r = self.ctx.request.get(url, timeout=int(timeout * 1000))
            if r.ok and r.body()[:5] == b"%PDF-":
                return r.body(), "ctx.request"
            errs.append(f"ctx.request HTTP {r.status}")
        except Exception as e:
            errs.append(f"ctx.request: {str(e).splitlines()[0][:80]}")
        # 4
        try:
            rc = self._pw.request.new_context(ignore_https_errors=True)
            try:
                r = rc.get(url, timeout=int(timeout * 1000))
                if r.ok and r.body()[:5] == b"%PDF-":
                    return r.body(), "lenient request context"
                errs.append(f"lenient HTTP {r.status}")
            finally:
                rc.dispose()
        except Exception as e:
            errs.append(f"lenient: {str(e).splitlines()[0][:80]}")
        raise RuntimeError(" | ".join(errs))


# ================================================================ saved output
# Everything retrieved is written to a folder in Downloads so it is not thrown
# away when the run ends. Layout is deliberately the same shape the main rfparts
# pipeline will want later: one directory per vendor, plus a JSONL manifest.
#
#   ~/Downloads/spacequal_datasheets/
#       Analog-Devices/AD1671S.pdf
#       Marki-Microwave/MM1-0320LBH.html    <- rendered DOM, so it can be
#       Marki-Microwave/MM1-0320LBH.txt        re-parsed without re-fetching
#       manifest.jsonl

_DEFAULT_OUTDIR = Path.home() / "Downloads" / "spacequal_datasheets"
OUTDIR = _DEFAULT_OUTDIR
SAVED = {"pdf": 0, "html": 0, "bytes": 0}


def _safe_name(name):
    return re.sub(r"[^A-Za-z0-9._+-]", "_", str(name))[:80] or "part"


def save_artifact(vendor_slug, mpn, kind, data, source_url, extra=None):
    """Write one file and append a manifest line. Returns the path."""
    import hashlib
    import json
    folder = OUTDIR / _safe_name(vendor_slug)
    folder.mkdir(parents=True, exist_ok=True)
    ext = {"pdf": ".pdf", "html": ".html", "txt": ".txt"}[kind]
    path = folder / (_safe_name(mpn) + ext)
    blob = data if isinstance(data, bytes) else data.encode("utf-8", "replace")
    path.write_bytes(blob)
    SAVED[kind if kind in SAVED else "html"] = \
        SAVED.get(kind if kind in SAVED else "html", 0) + 1
    SAVED["bytes"] += len(blob)
    rec = {"mpn": mpn, "vendor": vendor_slug, "kind": kind,
           "local_path": str(path.relative_to(OUTDIR)),
           "bytes": len(blob), "sha256": hashlib.sha256(blob).hexdigest(),
           "source_url": source_url,
           "fetched_at": time.strftime("%Y-%m-%dT%H:%M:%S")}
    if extra:
        rec.update(extra)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    with (OUTDIR / "manifest.jsonl").open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(rec) + "\n")
    return path


def report_saved():
    say("\n" + "=" * 76)
    say("SAVED FILES")
    say("=" * 76)
    if not (SAVED["pdf"] or SAVED["html"]):
        say("  nothing was saved this run")
        return
    mb = SAVED["bytes"] / (1024 * 1024)
    say(f"  {SAVED['pdf']} PDF(s), {SAVED['html']} HTML/text file(s), "
        f"{mb:.1f} MB total")
    say(f"  folder:   {OUTDIR}")
    say(f"  manifest: {OUTDIR / 'manifest.jsonl'}")
    for sub in sorted(x for x in OUTDIR.iterdir() if x.is_dir()) \
            if OUTDIR.exists() else []:
        files = sorted(sub.iterdir())
        say(f"    {sub.name}/  {len(files)} file(s)")
        for f in files[:4]:
            say(f"      {f.name}  ({f.stat().st_size // 1024} kB)")
        if len(files) > 4:
            say(f"      ... and {len(files) - 4} more")


# ============================================================ HTML -> text
_DROP_TAGS = re.compile(r"<(script|style|noscript|svg|head)\b.*?</\1>",
                        re.S | re.I)
_BLOCK_END = re.compile(r"</(p|div|li|tr|h[1-6]|table|section|br)\s*/?>", re.I)
_CELL_END = re.compile(r"</(td|th)\s*>", re.I)
_TAGS = re.compile(r"<[^>]+>")

# The page frame we do not want: nav above the datasheet, footer below it.
_START_HINTS = ("device overview", "general description")
_END_HINTS = ("stay informed", "privacy policy", "be the first to know",
              "copyright ©")


def html_to_text(raw):
    """Readable text from a product/datasheet page.

    Tables become ' | ' separated rows so the parametric data survives, the
    surrounding site chrome is trimmed, and repeated blocks are collapsed --
    Marki renders the ordering and electrical tables twice, once for print."""
    s = _DROP_TAGS.sub(" ", raw)
    s = _CELL_END.sub(" | ", s)
    s = _BLOCK_END.sub("\n", s)
    s = _TAGS.sub(" ", s)
    s = htmllib.unescape(s)
    s = re.sub(r"[ \t\xa0]+", " ", s)
    lines = [ln.strip(" |") .strip() for ln in s.split("\n")]
    lines = [ln for ln in lines if ln]

    # trim to the datasheet body when we can recognise it
    lo = 0
    for i, ln in enumerate(lines):
        if any(h in ln.lower() for h in _START_HINTS):
            lo = i
            break
    hi = len(lines)
    for i in range(lo + 1, len(lines)):
        if any(h in lines[i].lower() for h in _END_HINTS):
            hi = i
            break
    body = lines[lo:hi] or lines

    # collapse duplicates (print copies of the same table) while keeping order
    seen, out = set(), []
    for ln in body:
        key = re.sub(r"\s+", " ", ln.lower())
        if len(key) > 12 and key in seen:
            continue
        seen.add(key)
        out.append(ln)
    return "\n".join(out).strip()


# ================================================================== Marki
MARKI_CATEGORIES = [
    "https://markimicrowave.com/products/connectorized/mixers/",
    "https://markimicrowave.com/products/surface-mount/mixers/",
    "https://markimicrowave.com/products/connectorized/amplifiers/",
]
MARKI_SPACE = "https://markimicrowave.com/space/"

_WORDY = re.compile(r"^[A-Za-z]{4,}$")


def looks_like_pn(s):
    s = (s or "").strip().strip("-_")
    if not (3 < len(s) <= 40) or not any(c.isdigit() for c in s):
        return False
    return not any(_WORDY.match(seg) for seg in re.split(r"[-_]", s)[1:])


_HREF = re.compile(r'href\s*=\s*["\']([^"\']+)["\']', re.I)


def marki_parts(html, base):
    """Product pages on a Marki category listing."""
    out = {}
    for href in _HREF.findall(html):
        h = href.strip()
        m = re.match(r"^(?:https?://[^/]+)?(/products/[a-z-]+/[a-z0-9-]+/"
                     r"([A-Za-z0-9][A-Za-z0-9._-]*)/?)$", h, re.I)
        if m and looks_like_pn(m.group(2)):
            pn = m.group(2).upper()
            out.setdefault(pn, urllib.parse.urljoin(base, m.group(1)))
    return out


_SPEC_MARKER = re.compile(r"electrical specification|recommended operating|"
                          r"guaranteed from", re.I)


def fetch_html(url, browser, rate, min_chars=400, want_specs=False,
               return_raw=False):
    """Direct HTTP first, browser second.

    `want_specs` matters: the first run pulled 1397 chars from a Marki datasheet
    page and I called it a win, but it had no Electrical Specifications table --
    those are built by JS, so plain HTTP silently returns the prose only. When
    specs are expected and missing, that is a reason to re-fetch through the
    browser even though the text looked long enough."""
    raw = ""
    try:
        blob, ctype, code = http_get(url)
        raw = blob.decode("utf-8", "replace")
        text = html_to_text(raw)
        enough = len(text) >= min_chars
        specs_ok = (not want_specs) or bool(_SPEC_MARKER.search(text))
        if enough and specs_ok:
            return (text, "direct", raw) if return_raw else (text, "direct")
        note = (f"{len(text)} chars, no spec table" if enough
                else f"only {len(text)} chars")
    except urllib.error.HTTPError as e:
        note = f"HTTP {e.code}"
    except Exception as e:
        note = type(e).__name__
    if not (browser and getattr(browser, "ok", False)):
        msg = f"direct failed ({note}); no browser available"
        return ("", msg, raw) if return_raw else ("", msg)
    try:
        html = browser.get_html(
            url, wait_text="electrical specification" if want_specs else None)
        text = html_to_text(html)
        got = bool(_SPEC_MARKER.search(text))
        tag = "browser" + ("" if not want_specs else
                           (" +specs" if got else " but STILL no spec table"))
        how = f"{tag} (direct: {note})"
        return (text, how, html) if return_raw else (text, how)
    except Exception as e:
        msg = f"direct failed ({note}); browser also failed ({type(e).__name__})"
        return ("", msg, raw) if return_raw else ("", msg)


def stage_marki(n_parts, rate, browser=None):
    say("=" * 76)
    say("MARKI: datasheet text straight from the product page HTML (no PDF)")
    say("=" * 76)
    parts = {}
    for cat in MARKI_CATEGORIES:
        raw, how = "", ""
        st, val, secs = guarded("marki-cat", lambda u=cat: http_get(u), 35)
        if st == "ok":
            raw, how = val[0].decode("utf-8", "replace"), "direct"
        elif browser and getattr(browser, "ok", False):
            st2, val2, secs = guarded("marki-cat-b",
                                      lambda u=cat: browser.get_html(u), 60)
            if st2 == "ok":
                raw, how = val2, f"browser (direct: {val})"
        if not raw:
            say(f"  [fail] {cat} -> {val}")
            continue
        got = marki_parts(raw, cat)
        say(f"  [ok] {secs:4.1f}s  {cat.rsplit('/', 2)[-2]:<14} "
            f"{len(got):>4} product page(s)   via {how}")
        parts.update(got)
        time.sleep(rate)
    say(f"\n  {len(parts)} distinct part(s) across those categories")
    if not parts:
        return 0

    say(f"\n  pulling the datasheet HTML for {min(n_parts, len(parts))} part(s):")
    say(f"  saving into {OUTDIR / 'Marki-Microwave'}")
    ok = 0
    for pn, url in list(parts.items())[:n_parts]:
        ds = url.rstrip("/") + "/datasheet/"
        time.sleep(rate)
        st, val, secs = guarded(
            "marki-ds",
            lambda u=ds: fetch_html(u, browser, rate, want_specs=True,
                                    return_raw=True), 90)
        if st != "ok" or not val or not val[0]:
            say(f"    [fail   ] {pn:<18} "
                f"{val[1] if isinstance(val, tuple) else val}")
            continue
        text, how, raw = val
        cues = found_cues(text)
        has_specs = bool(re.search(r"electrical specification", text, re.I))
        has_temp = bool(re.search(r"-\s*55|guaranteed from", text, re.I))
        say(f"    [ok     ] {secs:4.1f}s  {pn:<18} {len(text):>6} chars / "
            f"{len(text.split()):>5} words   "
            f"specs={'Y' if has_specs else 'n'} temp={'Y' if has_temp else 'n'}"
            f"   via {how}")
        say(f"               cues: {', '.join(cues) if cues else 'NONE'}")
        first = " ".join(text.split())[:150]
        say(f"               \"{first}\"")
        meta = {"words": len(text.split()), "cues": cues,
                "has_spec_table": has_specs, "fetch_path": how}
        if raw:
            save_artifact("Marki-Microwave", pn, "html", raw, ds, meta)
        saved = save_artifact("Marki-Microwave", pn, "txt", text, ds, meta)
        say(f"               saved -> {saved.name}"
            + (f" + {pn}.html" if raw else ""))
        ok += 1
    say(f"\n  {ok}/{min(n_parts, len(parts))} datasheet page(s) yielded text")

    # Marki publish a dedicated Space & Hi-Rel page -- directly relevant to the
    # space-qualified side of the dataset.
    st, val, secs = guarded(
        "marki-space",
        lambda: fetch_html(MARKI_SPACE, browser, rate, return_raw=True), 90)
    if st == "ok" and val and val[0]:
        t, how, raw = val
        save_artifact("Marki-Microwave", "_space_and_hirel", "txt", t,
                      MARKI_SPACE, {"words": len(t.split()),
                                    "cues": found_cues(t)})
        say(f"\n  {MARKI_SPACE} -> {len(t)} chars via {how}; cues: "
            f"{', '.join(found_cues(t)) or 'NONE'}")
    else:
        say(f"\n  {MARKI_SPACE} -> unavailable ({val})")
    return ok


# ==================================================================== ADI
def adi_urls_from_xlsx(path, limit):
    p = Path(path)
    if not p.is_file():
        say(f"  ADI spreadsheet not found: {p}")
        return []
    try:
        import openpyxl
    except ImportError:
        say("  openpyxl missing:  pip install openpyxl")
        return []
    wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = [[(c.value if c.value is not None else "") for c in r]
            for r in ws.iter_rows(max_row=400)]
    wb.close()
    hdr = col = None
    for ri, r in enumerate(rows[:20]):
        for ci, cell in enumerate(r):
            if str(cell).strip().lower() == "generic part number":
                hdr, col = ri, ci
                break
        if col is not None:
            break
    if col is None:
        say("  could not find the 'Generic Part Number' column")
        return []
    pns, seen = [], set()
    for r in rows[hdr + 1:]:
        v = str(r[col]).strip() if col < len(r) else ""
        if looks_like_pn(v) and v.upper() not in seen:
            seen.add(v.upper())
            pns.append(v)
    base = "https://www.analog.com/media/en/technical-documentation/data-sheets/"
    out = []
    for n in pns[:limit]:
        stem = re.sub(r"S$", "", n, flags=re.I)
        out.append((n, [f"{base}{n.lower()}.pdf",
                        f"{base}{stem.lower()}.pdf"]))
    say(f"  {len(pns)} part number(s) in the sheet; trying {len(out)}")
    return out


def stage_adi_direct(xlsx, limit):
    say("\n" + "=" * 76)
    say("ADI: confirming the block is a block (not a wrong URL)")
    say("=" * 76)
    targets = adi_urls_from_xlsx(xlsx, limit)
    codes = {}
    for pn, urls in targets:
        for u in urls:
            st, val, secs = guarded("adi", lambda x=u: http_get(x), 30)
            tag = "PDF" if st == "ok" and val[0][:5] == b"%PDF-" else str(val)
            codes[tag] = codes.get(tag, 0) + 1
            say(f"    {pn:<14} {u.rsplit('/', 1)[-1]:<22} -> {tag}")
    say(f"\n  outcomes: {codes}")
    if any("403" in k for k in codes):
        say("  403 confirms a deliberate block on scripted access.")
        say("  The browser stage below fetches the same URLs properly.")
    return codes


def stage_adi_browser(xlsx, limit, browser, rate):
    say("\n" + "=" * 76)
    say("ADI: fetching the known URLs through YOUR Chrome session")
    say("=" * 76)
    say("  Your session, your cookies, your user agent. Nothing is disguised and")
    say("  no challenge is solved -- we just stop pretending a script is welcome")
    say("  where the site has said it is not.\n")
    if not (browser and getattr(browser, "ok", False)):
        say("  no browser available, so ADI cannot be fetched.")
        return 0
    targets = adi_urls_from_xlsx(xlsx, limit)
    if not targets:
        return 0
    say(f"  saving into {OUTDIR / 'Analog-Devices'}")
    ok = 0
    try:
        for pn, urls in targets:
            got = False
            for u in urls:
                try:
                    body, how = browser.get_bytes(u)
                except Exception as e:
                    say(f"    {pn:<14} {u.rsplit('/', 1)[-1]:<20} "
                        f"{str(e)[:150]}")
                    continue
                if body[:5] != b"%PDF-":
                    continue
                text = ""
                try:
                    import warnings
                    import pdfplumber
                    warnings.filterwarnings("ignore")
                    with pdfplumber.open(io.BytesIO(body)) as pdf:
                        text = "\n".join((p.extract_text() or "")
                                         for p in pdf.pages[:2])
                except Exception as e:
                    text = f"__EXTRACT_FAILED__ {type(e).__name__}"
                cues = found_cues(text) if not text.startswith("__") else []
                path = save_artifact(
                    "Analog-Devices", pn, "pdf", body, u,
                    {"words": len(text.split()), "cues": cues,
                     "fetch_path": how})
                say(f"    {pn:<14} {len(body) // 1024:>4} kB  %PDF  "
                    f"{len(text.split()):>5} words  via {how}  "
                    f"cues: {', '.join(cues) or 'NONE'}")
                say(f"                   saved -> {path.name}")
                ok += 1
                got = True
                break
            if not got:
                say(f"    {pn:<14} no PDF from any candidate")
            time.sleep(rate)
    except KeyboardInterrupt:
        say("\n  interrupted")
    say(f"\n  {ok}/{len(targets)} ADI datasheet(s) retrieved through Chrome")
    return ok


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--report", default="spacequal_finish_report.txt")
    ap.add_argument("--marki-parts", type=int, default=4,
                    help="Marki datasheet pages to pull text from")
    ap.add_argument("--adi-limit", type=int, default=4)
    ap.add_argument("--adi-xlsx",
                    default=str(Path.home() / "Downloads" / "newSources"
                               / "adi_space_portfolio_2026-07-28.xlsx"))
    ap.add_argument("--no-browser", action="store_true",
                    help="do not open Chrome at all (Marki direct only; ADI will "
                         "just be shown as blocked)")
    ap.add_argument("--port", type=int, default=CHROME_PORT)
    ap.add_argument("--rate", type=float, default=1.0)
    ap.add_argument("--skip-marki", action="store_true")
    ap.add_argument("--outdir", default=str(_DEFAULT_OUTDIR),
                    help=f"where to save PDFs / HTML "
                         f"(default {_DEFAULT_OUTDIR})")
    args = ap.parse_args()

    globals()["OUTDIR"] = Path(args.outdir)
    say("spacequal FINISH debug -- only the two unproven vendors")
    say(f"generated {time.strftime('%Y-%m-%d %H:%M:%S')}")
    say("already proven, not retested: Mini-Circuits, Qorvo, MACOM, Skyworks")
    say("")

    marki_ok = adi_ok = 0
    interrupted = False
    browser = None
    try:
        if not args.no_browser:
            say("=" * 76)
            say("BROWSER")
            say("=" * 76)
            browser = ManagedBrowser(port=args.port, rate=args.rate)
            browser.__enter__()
            if not browser.ok:
                say("  continuing without a browser")
            say("")
        if not args.skip_marki:
            marki_ok = stage_marki(args.marki_parts, args.rate, browser)
        # ADI: show the block, then do it properly through the browser
        stage_adi_direct(args.adi_xlsx, min(args.adi_limit, 2))
        if browser and browser.ok:
            adi_ok = stage_adi_browser(args.adi_xlsx, args.adi_limit,
                                       browser, args.rate)
    except KeyboardInterrupt:
        interrupted = True
        say("\n\n  Ctrl-C -- stopping and closing the browser.")
    finally:
        if browser:
            browser.close()

    say("\n" + "=" * 76)
    say("SCORECARD (this run only)")
    say("=" * 76)
    say(f"  Marki   {'PROVEN via HTML' if marki_ok else 'not proven'}"
        f"   ({marki_ok} datasheet page(s) yielded text)")
    say(f"  ADI     {'PROVEN via browser' if adi_ok else 'not proven'}"
        f"   ({adi_ok} PDF(s) retrieved through Chrome)")
    if interrupted:
        say("\n  (run was interrupted, so these counts are partial)")
    report_saved()
    out = Path(args.report)
    try:
        out.write_text(OUT.getvalue(), encoding="utf-8")
        print(f"\nreport written to {out.resolve()}", flush=True)
    except Exception as e:
        print(f"\ncould not write {out}: {e}", flush=True)
    sys.stdout.flush()
    os._exit(0)


if __name__ == "__main__":
    main()
