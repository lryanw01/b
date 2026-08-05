"""dbg_part_specs.py — one part number in, every source's raw data out.

    python dbg_part_specs.py HMC8038
    python dbg_part_specs.py MABA-011016 SKY53759-11
    python dbg_part_specs.py HMC8038 --key freq
    python dbg_part_specs.py HMC8038 --no-scan        (database only, fast)

WHAT IT DOES
    1. Finds the part in the database: every stored spec row with the source URL,
       method and confidence that produced it, then what each key resolves to
       after the confidence contest -- which is what the grid and ranker see.
    2. Works out which SOURCES the part came from, both from those stored rows and
       by searching the source trees on disk.
    3. Dumps the RAW record from each source in that source's own shape:

         ADI / any workbook   the spreadsheet row, cell by cell, with the
                              frequency conversion the ingest would apply and
                              WHY (cell unit / header unit / magnitude guess)
         everythingRF         the parsed attribute rows for that product box
         MACOM                the data-part JSON, including unmapped fields
         Qorvo                the parametric grid row and its section
         Marki                sections, ordering options, inline specs
         local datasheet      extracted text, sniffed type, and exactly which
                              PARAM_SPECS patterns fire on it
         anything else        the raw text around the part number

WHY
    A wrong number has three possible homes needing different fixes: the source
    says something unexpected, the parser mis-reads it, or a stale row from an
    older parser is still winning the confidence contest. Only the stored rows and
    the raw source side by side tell them apart.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

try:
    from pythonrfparts import partdb
except Exception as e:                                   # pragma: no cover
    sys.exit(f"Could not import the package: {e}\n"
             "Run this from the folder that contains pythonrfparts/.")


def _opt(name):
    """Import a package module, or None -- every source reader is optional so a
    missing dependency degrades one section instead of the whole script."""
    try:
        return __import__(f"pythonrfparts.{name}", fromlist=[name])
    except Exception:
        return None


dsmine = _opt("dsmine")
adi_par = _opt("adi_parametric_ingest")
esi = _opt("erf_space_ingest")
vc = _opt("vendor_catalogs")

_TEXT_EXT = {".html", ".htm", ".txt", ".csv", ".json"}
_BOOK_EXT = {".xlsx", ".xlsm"}
_FREQ_HDR = re.compile(r"freq|bandwidth|cutoff", re.I)


def loose(s):
    return re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


# ----------------------------------------------------------------- database
def show_db(mpn, only_key=None):
    conn = partdb.db()
    rows = conn.execute(
        "SELECT id, mpn, vendor, category, subcategory, description, product_url "
        "FROM parts WHERE upper(mpn)=? OR upper(mpn) LIKE ?",
        (mpn.upper(), f"%{mpn.upper()}%")).fetchall()
    if not rows:
        print("  not in the database. Still scanning the sources below: a part "
              "can be present in a source and fail to ingest.")
        return set()
    hints = set()
    for r in rows:
        print(f"\n  {'=' * 68}")
        print(f"  {r['mpn']}   vendor={r['vendor']}   "
              f"category={r['category']}/{r['subcategory'] or '-'}")
        if r["description"]:
            print(f"  {r['description'][:96]}")
        if r["product_url"]:
            print(f"  {r['product_url'][:96]}")
        specs = conn.execute(
            "SELECT key, value_min, value_typ, value_max, unit, method, "
            "confidence, source_url, snippet FROM specs WHERE part_id=? "
            "ORDER BY key, confidence DESC", (r["id"],)).fetchall()
        if only_key:
            specs = [s for s in specs if only_key.lower() in s["key"].lower()]
        print(f"\n  {len(specs)} stored spec row(s):")
        print(f"    {'key':<22}{'min':>10}{'typ':>10}{'max':>10} {'unit':<6}"
              f"{'conf':>5}  method / source")
        for s in specs:
            def f(v):
                return "" if v is None else f"{v:g}"
            print(f"    {s['key']:<22}{f(s['value_min']):>10}"
                  f"{f(s['value_typ']):>10}{f(s['value_max']):>10} "
                  f"{(s['unit'] or ''):<6}{s['confidence']:>5.2f}  "
                  f"{s['method'] or '?'} / {(s['source_url'] or '-')[:40]}")
            if s["snippet"]:
                print(f"      snippet: {str(s['snippet'])[:86]}")
            for token in (s["method"], s["source_url"], s["snippet"]):
                if token:
                    hints.add(str(token)[:150])
        ev = conn.execute("SELECT signal, weight, source_url FROM qual_evidence "
                          "WHERE part_id=?", (r["id"],)).fetchall()
        if ev:
            print(f"\n  {len(ev)} evidence row(s):")
            for e in ev:
                print(f"    {e['signal']:<34} w={e['weight']}  "
                      f"{(e['source_url'] or '')[:38]}")
                hints.add(str(e["signal"]))
        cands = [c for c in partdb.query_candidates(limit=100000)
                 if loose(c.get("model")) == loose(r["mpn"])]
        if cands:
            resolved = cands[0].get("specs") or {}
            print("\n  resolved for the grid/ranker:")
            for k in sorted(resolved):
                if only_key and only_key.lower() not in k.lower():
                    continue
                print(f"    {k:<26} {resolved[k]}")
            _sanity(resolved, r["description"] or "")
        else:
            print("\n  ! in `parts` but query_candidates does not return it "
                  "(a category or frequency filter is excluding it)")
    return hints


_DESC_RANGE = re.compile(
    r"(\d+(?:\.\d+)?)\s*(ghz|mhz|khz)?\s*(?:to|-|\u2013|through)\s*"
    r"(\d+(?:\.\d+)?)\s*(ghz|mhz|khz)", re.I)
_DESC_SCALE = {"ghz": 1.0, "mhz": 1e-3, "khz": 1e-6, "": 1.0}


def band_from_text(text):
    """A frequency range stated in prose, in GHz, or None.

    Datasheet titles and descriptions usually spell the band out ("0.1 GHz to
    6.0 GHz"), which is an independent check on the parsed number. A stored band
    that contradicts the part's own description is wrong even when it looks
    physically plausible -- which is the case a magnitude sanity test cannot
    catch on its own."""
    m = _DESC_RANGE.search(str(text or ""))
    if not m:
        return None
    hi_u = (m.group(4) or "ghz").lower()
    lo_u = (m.group(2) or hi_u).lower()
    try:
        lo = float(m.group(1)) * _DESC_SCALE.get(lo_u, 1.0)
        hi = float(m.group(3)) * _DESC_SCALE.get(hi_u, 1.0)
    except ValueError:
        return None
    return (min(lo, hi), max(lo, hi))


def _sanity(resolved, description=""):
    fg = resolved.get("freq_ghz")
    if isinstance(fg, (tuple, list)) and len(fg) == 2:
        lo, hi = fg
        note = "looks sane"
        if hi > 300:
            note = "!! above 300 GHz -- almost certainly a unit error"
        elif lo > 0 and hi / lo > 1000:
            note = "!! spans >3 decades -- almost certainly a unit error"
        print(f"\n  frequency band {lo} - {hi} GHz -> {note}")
        if hi > 300:
            print(f"     if the source meant MHz, the band is "
                  f"{lo / 1000:g} - {hi / 1000:g} GHz")
        stated = band_from_text(description)
        if stated:
            slo, shi = stated
            agree = (abs(slo - lo) <= max(0.05, slo * 0.05)
                     and abs(shi - hi) <= max(0.05, shi * 0.05))
            print(f"  the part's own description says {slo:g} - {shi:g} GHz "
                  f"-> {'agrees' if agree else '!! DISAGREES with the stored band'}")
            if not agree:
                print("     the description is written by the vendor, so trust it "
                      "over the parsed\n     value: this part needs its source "
                      "re-parsed.")


def guess_sources(hints):
    blob = " ".join(hints).lower()
    out = []
    for name, needles in (
            ("everythingRF", ("everythingrf", "erf-space", "aggregator")),
            ("ADI parametrics", ("analog.com", "parametric")),
            ("ADI space list", ("adi-space", "qml", "class s", "class-s")),
            ("MACOM", ("macom", "data-part")),
            ("Marki", ("marki",)),
            ("Qorvo", ("qorvo",)),
            ("Skyworks", ("skyworks",)),
            ("TI", ("ti-space", "texas instruments", "ti.com")),
            ("local datasheet mining", ("mined from", "datasheet")),
            ("switch-type hotfix", ("hotfix-switch-types",))):
        if any(n in blob for n in needles):
            out.append(name)
    return out


# ------------------------------------------------------------- source trees
def source_roots():
    roots = []
    try:
        paths = __import__("pythonrfparts.paths", fromlist=["paths"])
        for attr in ("SOURCES_ROOT", "EVERYTHING_RF", "NEW_SOURCES",
                     "QORVO_PAGES", "DATASHEET_DIR", "VENDOR_CACHE",
                     "DATA_ROOT"):
            v = getattr(paths, attr, None)
            if v:
                roots.append(Path(v))
    except Exception:
        pass
    roots += [Path(partdb.DATA), Path(partdb.DATA).parent,
              Path.home() / "Downloads" / "rfparts"]
    seen, out = set(), []
    for r in roots:
        try:
            rp = r.resolve()
        except OSError:
            continue
        if rp.is_dir() and rp not in seen:
            # skip a root already covered by a shorter one
            if any(str(rp).startswith(str(s) + "/") or
                   str(rp).startswith(str(s) + "\\") for s in seen):
                continue
            seen.add(rp)
            out.append(rp)
    return out


def find_files(mpn, limit=25, max_scanned=60000):
    """Files in the source trees that mention this part, classified by shape."""
    target = loose(mpn)
    hits, scanned, books = [], 0, []
    for root in source_roots():
        try:
            walker = root.rglob("*")
        except OSError:
            continue
        for f in walker:
            if len(hits) >= limit or scanned >= max_scanned:
                break
            try:
                if not f.is_file():
                    continue
            except OSError:
                continue
            ext = f.suffix.lower()
            if ext in _BOOK_EXT:
                if not f.name.startswith("~$"):
                    books.append(f)
                continue
            if ext == ".pdf":
                if target and target in loose(f.stem):
                    hits.append((f, "datasheet"))
                continue
            if ext not in _TEXT_EXT:
                continue
            scanned += 1
            try:
                blob = f.read_text(encoding="utf-8", errors="replace")
            except (OSError, MemoryError):
                continue
            if target and target in loose(f.stem):
                hits.append((f, "datasheet"))
            elif target and target in loose(blob):
                hits.append((f, _classify(blob, f)))
    # workbooks are checked by opening them, not by text search
    for b in books[:12]:
        hits.append((b, "workbook"))
    # The source roots nest (Sources lives inside the repo root), so the same
    # file arrives more than once. Dedupe on the resolved path.
    seen, unique = set(), []
    for f, kind in hits:
        try:
            key = f.resolve()
        except OSError:
            key = f
        if key in seen:
            continue
        seen.add(key)
        unique.append((f, kind))
    return unique, scanned


def _classify(blob, path):
    head = blob[:400000]
    low = head.lower()
    if "data-part" in head and "macom" in low:
        return "macom"
    if "product-box" in head or "manu-name" in head:
        return "everythingrf"
    if "grid-row" in head and "product-name" in head:
        return "qorvo"
    if "markimicrowave" in low or "general-description" in head:
        return "marki"
    return "generic"


# --------------------------------------------------------------- per-source
def dump_workbook(path, mpn):
    try:
        import openpyxl
    except ImportError:
        print("      (openpyxl not installed; cannot read workbooks)")
        return False
    target = loose(mpn)
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=False)
    except Exception:
        return False
    shown = False
    try:
        for sheet in wb.sheetnames:
            rows = []
            for i, row in enumerate(wb[sheet].iter_rows(values_only=True)):
                rows.append(row)
                if i > 6000:
                    break
            hdr_i = 0
            for i, row in enumerate(rows[:14]):
                if row and any(isinstance(c, str) and "part" in c.lower()
                               for c in row if c):
                    hdr_i = i
                    break
            headers = [str(c or "") for c in (rows[hdr_i] if rows else [])]
            for row in rows[hdr_i + 1:]:
                if not row:
                    continue
                if target and target in loose(" ".join(str(c) for c in row if c)):
                    print(f"\n      {path.name} [{sheet}] raw row:")
                    for h, v in zip(headers, row):
                        if v is None or str(v).strip() == "":
                            continue
                        line = f"        {h[:38]:<38} = {v!r}"
                        if adi_par and _FREQ_HDR.search(h):
                            scale = adi_par.freq_unit_from_header(h)
                            cell = None
                            if hasattr(adi_par, "freq_from_cell"):
                                cell = adi_par.freq_from_cell(v)
                            conv = adi_par._freq_to_ghz(v, scale)
                            why = ("cell unit" if cell else
                                   "header unit" if scale else "magnitude guess")
                            line += f"  -> {conv} GHz [{why}]"
                        print(line)
                    shown = True
                    break
            if shown:
                break
    finally:
        wb.close()
    return shown


def dump_everythingrf(path, mpn):
    if not esi:
        return False
    target = loose(mpn)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    try:
        parts = list(esi.parse_page(text, path.parent.name))
    except Exception as e:
        print(f"      ({path.name}: erf parse failed, {type(e).__name__})")
        return False
    for part in parts:
        if loose(part.get("mpn")) != target:
            continue
        print(f"\n      {path.name} (folder {path.parent.name}) — everythingRF")
        print(f"        title       : {str(part.get('title', ''))[:74]}")
        print(f"        description : {str(part.get('description', ''))[:74]}")
        print(f"        category    : {part.get('category')}/"
              f"{part.get('subcategory') or '-'}")
        print(f"        grade text  : {part.get('grade_text', '')!r}")
        print("        parsed spec rows:")
        for r in part.get("spec_rows") or []:
            print(f"          {r.key:<20} min={r.value_min} typ={r.value_typ} "
                  f"max={r.value_max} unit={r.unit!r} conf={r.confidence}")
        return True
    return False


def dump_macom(path, mpn):
    if not vc or not hasattr(vc, "parse_macom_data_parts"):
        return False
    target = loose(mpn)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    try:
        recs = vc.parse_macom_data_parts(text, path.name)
    except Exception as e:
        print(f"      ({path.name}: macom parse failed, {type(e).__name__})")
        return False
    for pn, rec in recs.items():
        if loose(pn) != target:
            continue
        print(f"\n      {path.name} — MACOM data-part record")
        print(f"        description : {rec.get('description', '')[:72]}")
        print(f"        datasheet   : {rec.get('datasheet_url') or '(none)'}")
        print("        parsed specs:")
        for k, v in sorted(rec.get("specs", {}).items()):
            print(f"          {k:<22} {v}")
        m = re.search(r'data-part\s*=\s*"(\{.*?\})"\s*>', text, re.S)
        blobs = re.findall(r'data-part\s*=\s*"(\{.*?\})"\s*>', text, re.S)
        import html as _h
        for b in blobs:
            if target not in loose(b):
                continue
            try:
                raw = json.loads(_h.unescape(b), strict=False)
            except Exception:
                continue
            print(f"        RAW json fields: {sorted(raw)}")
            for sp in (raw.get("specs") or [])[:16]:
                print(f"          raw spec: {sp}")
            for at in (raw.get("attributes") or [])[:10]:
                print(f"          raw attr: {at}")
            break
        return True
    return False


def dump_qorvo(path, mpn):
    if not vc or not hasattr(vc, "parse_qorvo_parametric"):
        return False
    target = loose(mpn)
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    try:
        recs = vc.parse_qorvo_parametric(text, "", "", path.name)
    except Exception as e:
        print(f"      ({path.name}: qorvo parse failed, {type(e).__name__})")
        return False
    for pn, rec in recs.items():
        if loose(pn) != target:
            continue
        print(f"\n      {path.name} — Qorvo parametric row")
        print(f"        section     : {rec.get('section', '')}")
        print(f"        category    : {rec.get('category')}/"
              f"{rec.get('subcategory') or '-'}")
        print(f"        description : {rec.get('description', '')[:70]}")
        print(f"        datasheet   : {rec.get('datasheet_url') or '(none)'}")
        for k, v in sorted(rec.get("specs", {}).items()):
            print(f"          {k:<22} {v}")
        return True
    return False


def dump_marki(path, mpn):
    if not vc or not hasattr(vc, "marki_datasheet_text"):
        return False
    try:
        raw = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    if loose(mpn) not in loose(raw):
        return False
    text = vc.marki_datasheet_text(raw)
    rec = {"mpn": mpn, "specs": {}}
    try:
        vc._marki_specs_from_text(rec, text)
    except Exception:
        pass
    print(f"\n      {path.name} — Marki page")
    if hasattr(vc, "marki_order_options"):
        print(f"        ordering options: {vc.marki_order_options(text)}")
    if hasattr(vc, "marki_inline_specs"):
        print(f"        inline specs    : {vc.marki_inline_specs(text)}")
    print("        parsed specs:")
    for k, v in sorted(rec["specs"].items()):
        print(f"          {k:<22} {v}")
    return True


def dump_datasheet(path, mpn):
    """What the enrichment step actually sees. The place to look when mining
    yields nothing despite the text extracting fine."""
    if not dsmine:
        return False
    kind = dsmine._sniff(path)
    text = dsmine.datasheet_text(path)
    mined = dsmine.mine_text(text)
    print(f"\n      {path.name} — local datasheet")
    print(f"        sniffed type : {kind}")
    print(f"        text chars   : {len(text)}")
    print(f"        patterns hit : {sorted(mined) or 'NONE'}")
    for k, v in sorted(mined.items()):
        print(f"          {k:<22} {v}")
    if text:
        snippet = re.sub(r"\s+", " ", text[:400])
        print(f"        text starts  : {snippet}")
        if not mined:
            print("        ! text extracted but NO pattern matched: the wording "
                  "differs from what")
            print("          PARAM_SPECS expects. The snippet above is what a "
                  "new pattern must match.")
    elif kind == "corrupt":
        print("        ! file rejected as corrupt (text-mangled download) -- "
              "re-download in binary mode")
    return True


def dump_generic(path, mpn):
    try:
        blob = path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return False
    i = blob.upper().find(mpn.upper())
    if i < 0:
        return False
    ctx = re.sub(r"\s+", " ", blob[max(0, i - 180):i + 260])
    print(f"\n      {path.name} — raw context")
    print(f"        ...{ctx}...")
    return True


DUMPERS = {"workbook": dump_workbook, "everythingrf": dump_everythingrf,
           "macom": dump_macom, "qorvo": dump_qorvo, "marki": dump_marki,
           "datasheet": dump_datasheet, "generic": dump_generic}


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__.split("\n")[0])
    ap.add_argument("mpn", nargs="+")
    ap.add_argument("--no-scan", action="store_true",
                    help="database only; do not read source files")
    ap.add_argument("--files", type=int, default=25,
                    help="max source files to dump per part (default 25)")
    ap.add_argument("--key", default=None,
                    help="only show spec keys containing this")
    args = ap.parse_args(argv)

    print(f"database     : {partdb.DB_PATH}")
    for r in source_roots():
        print(f"source tree  : {r}")
    for mpn in args.mpn:
        print(f"\n{'#' * 72}\n#  {mpn}\n{'#' * 72}")
        hints = show_db(mpn, args.key)
        guessed = guess_sources(hints)
        if guessed:
            print(f"\n  sources implied by the stored rows: {', '.join(guessed)}")
        if args.no_scan:
            continue
        print("\n  --- raw data from the sources on disk ---")
        hits, scanned = find_files(mpn, limit=args.files)
        print(f"  {scanned} text file(s) examined, {len(hits)} candidate(s)")
        shown = 0
        for path, kind in hits:
            fn = DUMPERS.get(kind)
            try:
                ok = bool(fn(path, mpn)) if fn else False
                if not ok and kind not in ("workbook", "generic"):
                    ok = dump_generic(path, mpn)
            except Exception as e:
                print(f"      {path.name}: dump failed "
                      f"({type(e).__name__}: {e})")
                ok = False
            shown += 1 if ok else 0
        if not shown:
            print("  no source file yielded a record for this part.")
    print("\nHow to read it: if a source's RAW value is right but the stored row")
    print("is wrong, the parser mis-scaled it. If the raw value is already odd,")
    print("the source says so. If an old high-confidence row beats a newer")
    print("correct one, re-parse that source (bump PARSER_VERSION or touch the")
    print("file). If a datasheet extracts text but hits NO patterns, the wording")
    print("differs from PARAM_SPECS and the snippet shows what to add.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
